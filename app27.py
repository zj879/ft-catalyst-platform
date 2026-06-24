"""
费托合成催化剂材料数据平台
论文自动下载模块：DOI分类 → 自动下载 → PDF校验（manifest）→ 导出结果
"""

import streamlit as st
import pandas as pd
import json
import hashlib
import re
import time
import subprocess
import sys
import os
import plotly.express as px
import plotly.graph_objects as go

try:
    import statsmodels.api  # noqa: F401  仅用于检测是否安装，供散点图趋势线功能判断
    _HAS_STATSMODELS = True
except ImportError:
    _HAS_STATSMODELS = False
from typing import Optional
from datetime import datetime, timezone
from collections import Counter
from pathlib import Path

# ══════════════════════════════════════════════════════════════════
#  页面配置
# ══════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="费托合成催化剂材料数据平台",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ══════════════════════════════════════════════════════════════════
#  全局样式
# ══════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;600&display=swap');
html, body, [class*="css"] {
    font-family: 'Times New Roman', 'Microsoft YaHei', '微软雅黑', serif !important;
    font-size: 17px !important;
}
p, div, span, li, td, th, label, button, input, textarea, select,
h1, h2, h3, h4, h5, h6, .stMarkdown, .stText {
    font-family: 'Times New Roman', 'Microsoft YaHei', '微软雅黑', serif !important;
}
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 0rem !important; padding-bottom: 2rem !important; }

/* 隐藏侧边栏及展开按钮 */
[data-testid="stSidebar"] { display: none !important; }
[data-testid="collapsedControl"] { display: none !important; }

/* 顶部导航栏 */
.top-navbar {
    background: linear-gradient(135deg, #0D2B5E 0%, #1565C0 100%);
    padding: 0;
    margin-bottom: 0;
    box-shadow: 0 2px 8px rgba(0,0,0,0.25);
}
.top-navbar-header {
    display: flex;
    align-items: center;
    gap: 16px;
    padding: 14px 32px 10px 24px;
}
.top-navbar-header img {
    height: 52px;
    width: auto;
    border-radius: 6px;
    flex-shrink: 0;
}
.top-navbar-title {
    color: #fff;
    font-size: 1.35rem;
    font-weight: 700;
    line-height: 1.25;
    letter-spacing: -0.01em;
}
.top-navbar-subtitle {
    color: rgba(255,255,255,0.75);
    font-size: 0.78rem;
    margin-top: 2px;
}
.top-navbar-tabs {
    display: flex;
    gap: 2px;
    padding: 0 24px;
    border-top: 1px solid rgba(255,255,255,0.12);
}
.nav-tab {
    padding: 10px 22px;
    color: rgba(255,255,255,0.75) !important;
    font-size: 0.92rem;
    font-weight: 500;
    cursor: pointer;
    border-bottom: 3px solid transparent;
    transition: all 0.18s;
    text-decoration: none;
    white-space: nowrap;
    background: none;
    border-left: none;
    border-right: none;
    border-top: none;
}
.nav-tab:hover { color: #fff !important; background: rgba(255,255,255,0.08); }
.nav-tab.active {
    color: #fff !important;
    border-bottom: 3px solid #fff;
    font-weight: 700;
    background: rgba(255,255,255,0.1);
}

/* Hero */
.hero-banner {
    background: linear-gradient(135deg,#0D2B5E 0%,#1565C0 40%,#1976D2 100%);
    border-radius:10px; padding:26px 32px 22px; margin-bottom:20px;
    position:relative; overflow:hidden;
}
.hero-banner::before {
    content:''; position:absolute; top:-40px; right:-40px;
    width:200px; height:200px; background:rgba(255,255,255,0.05); border-radius:50%;
}
.hero-tag {
    display:inline-block; background:rgba(255,255,255,0.18);
    border:1px solid rgba(255,255,255,0.3); color:#FFE0E0;
    font-size:0.75rem; padding:3px 12px; border-radius:20px;
    margin-bottom:12px; letter-spacing:0.04em;
}
.hero-title { font-size:2.0rem; font-weight:700; color:#fff; margin:0 0 8px; letter-spacing:-0.02em; line-height:1.2; }
.hero-sub   { font-size:0.88rem; color:rgba(255,255,255,0.8); margin:0; line-height:1.5; }

/* 统计卡片 */
.stat-card { background:#fff; border:1px solid #E8E8E8; border-radius:8px; padding:18px 20px; box-shadow:0 1px 4px rgba(0,0,0,0.06); }
.stat-label { font-size:0.92rem; color:#888; margin-bottom:6px; }
.stat-value { font-size:2.1rem; font-weight:700; color:#1A1A2E; line-height:1; font-family:'JetBrains Mono',monospace; }
.stat-desc  { font-size:0.75rem; color:#AAA; margin-top:4px; }

/* 区块标题 */
.section-title { font-size:1.1rem; font-weight:600; color:#0D2B5E; border-left:3px solid #1565C0; padding-left:10px; margin:22px 0 12px; }

/* 提示框 */
.info-box { background:#FFF8F8; border:1px solid #BBDEFB; border-radius:8px; padding:12px 16px; font-size:0.83rem; line-height:1.65; color:#4A0000; }
.info-box code { background:#E3F2FD; color:#1565C0; padding:1px 5px; border-radius:3px; font-family:'JetBrains Mono',monospace; font-size:0.8rem; }

/* 流程步骤条 */
.pipeline-bar {
    display:flex; align-items:center; gap:0;
    background:#F8F8F8; border:1px solid #EEE;
    border-radius:10px; padding:4px; margin-bottom:24px;
}
.pip-step {
    flex:1; text-align:center; padding:10px 6px; border-radius:7px;
    font-size:0.8rem; font-weight:500; cursor:default; transition:all .2s;
}
.pip-step.done    { background:#E3F2FD; color:#1565C0; }
.pip-step.active  { background:#1565C0; color:#fff; font-weight:700; box-shadow:0 2px 8px rgba(21,101,192,0.3); }
.pip-step.pending { background:transparent; color:#BBB; }
.pip-arrow { color:#DDD; font-size:0.9rem; padding:0 2px; }

/* 结果行 */
.result-row {
    display:flex; align-items:center; gap:12px;
    padding:8px 14px; border-radius:6px; margin-bottom:5px; font-size:0.83rem;
}
.result-ok   { background:#F0FFF4; border-left:3px solid #4CAF50; }
.result-warn { background:#FFF8E1; border-left:3px solid #FF9800; }
.result-err  { background:#E3F2FD; border-left:3px solid #1565C0; }

/* 修复 expander 箭头图标在字体未加载时显示为 keyboard_arrow_down 文字 */
[data-testid="stExpander"] summary svg { display: inline !important; }
[data-testid="stExpander"] summary span[data-testid="stExpanderToggleIcon"] {
    font-size: 0 !important;
    line-height: 0 !important;
}
[data-testid="stExpander"] summary span[data-testid="stExpanderToggleIcon"] svg {
    font-size: initial !important;
    width: 1.2rem !important;
    height: 1.2rem !important;
}

/* 主按钮 */
.stButton > button[kind="primary"] {
    background:linear-gradient(135deg,#1565C0,#1976D2) !important;
    border:none !important; color:#fff !important;
    font-weight:600 !important; border-radius:6px !important;
}
.stButton > button[kind="primary"]:hover {
    background:linear-gradient(135deg,#0D2B5E,#0D47A1) !important;
    box-shadow:0 2px 8px rgba(21,101,192,0.35) !important;
}

/* 彻底隐藏上传组件内的按钮，只保留拖拽区域 */
[data-testid="stFileUploaderDropzone"] button {
    display: none !important;
}
[data-testid="stFileUploaderDropzone"] small {
    display: none !important;
}
[data-testid="stFileUploaderDropzone"] {
    justify-content: center !important;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
#  核心常量与函数
# ══════════════════════════════════════════════════════════════════

PUBLISHER_ROUTES = {
    "10.1002": ("Wiley",            "webvpn",   "https://onlinelibrary.wiley.com/doi/pdfdirect/{doi}"),
    "10.1021": ("ACS",              "verified", "https://pubs.acs.org/doi/pdf/{doi}"),
    "10.1007": ("Springer Nature",  "verified", "https://link.springer.com/content/pdf/{doi}.pdf"),
    "10.1016": ("Elsevier",         "webvpn",   None),
    "10.1039": ("RSC",              "webvpn",   None),
    "10.3390": ("MDPI",             "direct",   "https://www.mdpi.com/{doi}/pdf"),
    "10.1080": ("Taylor & Francis", "probe",    None),
    "10.1038": ("Nature",           "direct",   "https://www.nature.com/articles/{suffix}.pdf"),
    "10.1126": ("Science",          "webvpn",   None),
    "10.1103": ("APS",              "direct",   "https://journals.aps.org/doi/{doi}"),
    "10.1149": ("IOP/ECS",          "webvpn",   None),
    "10.1515": ("De Gruyter",       "probe",    None),
    "10.1590": ("SciELO",           "direct",   None),
}

ROUTE_LABEL = {
    "direct":   "🟢 直接下载",
    "verified": "🔵 已验证路由",
    "webvpn":   "🟠 需 WebVPN",
    "probe":    "🟡 出版商探针",
    "review":   "🔴 人工审核",
    "invalid":  "⚫ 无效 DOI",
}

AUTO_ROUTES = {"direct", "verified"}
PUBLISHER_DELAY = {"ACS": 3.0, "Springer Nature": 2.0, "MDPI": 1.5, "Nature": 2.0, "APS": 2.0}
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "application/pdf,application/octet-stream,*/*",
}


def classify_doi(doi: str, access_type: str = "") -> dict:
    doi = doi.strip()
    if not doi or doi.lower() in ("nan", "none", ""):
        return {"route": "invalid", "publisher": "未知", "url": ""}
    prefix = ".".join(doi.split("/")[0].split(".")[:2]) if "/" in doi else ""
    if prefix not in PUBLISHER_ROUTES:
        return {"route": "review", "publisher": "低频/未知", "url": ""}
    pub, route, tpl = PUBLISHER_ROUTES[prefix]
    if str(access_type).strip().upper() in ("OA", "OA_DOWNLOADABLE", "开放获取") and route == "webvpn":
        route = "direct"
    url = ""
    if tpl:
        suffix = doi.split("/")[-1] if "/" in doi else doi
        url = tpl.replace("{doi}", doi).replace("{suffix}", suffix)
    return {"route": route, "publisher": pub, "url": url}


def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    mapping = {}
    for c in df.columns:
        lc = c.strip().lower().replace(" ", "_").replace("-", "_")
        if (lc == "doi" or lc.startswith("doi")) and "doi" not in mapping.values():
            mapping[c] = "doi"
        elif any(k in lc for k in ("access", "_oa", "classification")) and "access_type" not in mapping.values():
            mapping[c] = "access_type"
        elif "title" in lc and "title" not in mapping.values():
            mapping[c] = "title"
        elif "year" in lc and "year" not in mapping.values():
            mapping[c] = "year"
    return df.rename(columns=mapping)


def classify_df(df: pd.DataFrame) -> tuple[pd.DataFrame, bool]:
    df = normalize_cols(df.copy())
    if "doi" not in df.columns:
        st.error(f"❌ 未找到 DOI 列。当前列名：**{'、'.join(df.columns.tolist()[:8])}**")
        return df, False
    access = df.get("access_type", pd.Series([""] * len(df), index=df.index))
    results = [classify_doi(str(df["doi"].iloc[i]), str(access.iloc[i]) if i < len(access) else "")
               for i in range(len(df))]
    df["路由"]    = [r["route"]     for r in results]
    df["出版商"]  = [r["publisher"] for r in results]
    df["下载URL"] = [r["url"]       for r in results]
    df["状态"]    = "待下载"
    return df, True


def load_df(file) -> pd.DataFrame:
    name = file.name.lower()
    if name.endswith(".csv"):
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                file.seek(0)
                return pd.read_csv(file, encoding=enc)
            except Exception:
                continue
    return pd.read_excel(file, engine="openpyxl")


def safe_filename(doi: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", doi) + ".pdf"


def validate_pdf_bytes(data: bytes, min_kb: int = 20) -> dict:
    ok_header = data[:5] == b"%PDF-"
    ok_size   = len(data) > min_kb * 1024
    sha       = hashlib.sha256(data).hexdigest()
    return {
        "valid":   ok_header and ok_size,
        "header":  ok_header,
        "size_kb": round(len(data) / 1024, 1),
        "sha256":  sha,
        "reason":  ("" if (ok_header and ok_size) else
                    ("非PDF内容" if not ok_header else f"文件过小({len(data)//1024}KB)")),
    }


def _parse_filename(path: Path) -> dict:
    """从文件名猜测作者/年份/标题，来自 build_manifest.py。"""
    stem = path.stem
    match = re.search(r"_([0-9a-f]{10,12})$", stem, flags=re.I)
    hash_suffix = ""
    if match:
        hash_suffix = match.group(1).lower()
        stem = stem[: match.start()]
    year = author = ""
    title = stem.replace("_", " ")
    ym = re.search(r"(^|_)(19|20)\d{2}(_|$)", stem)
    if ym:
        year   = ym.group(0).strip("_")
        before = stem[:ym.start()].strip("_")
        after  = stem[ym.end():].strip("_")
        author = before.replace("_", " ").strip()
        title  = after.replace("_", " ").strip() or title
    return {"filename_author_guess": author, "filename_year_guess": year,
            "filename_title_guess": title,   "filename_hash_suffix": hash_suffix}


def _sha256_file(path: Path) -> str:
    """计算文件 SHA-256，来自 build_manifest.py。"""
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_pdf_info(path: Path) -> dict:
    """读取 PDF 元信息，完整复用 build_manifest.py 的 read_pdf_info。"""
    info = {
        "pdf_header_valid":  False,
        "pdf_read_ok":       False,
        "pdf_error":         "",
        "page_count":        "",
        "is_encrypted":      "",
        "pdf_metadata_title":  "",
        "pdf_metadata_author": "",
        "first_page_doi":    "",
    }
    try:
        with path.open("rb") as f:
            info["pdf_header_valid"] = (f.read(5) == b"%PDF-")
    except OSError as e:
        info["pdf_error"] = f"header_read_error: {e}"
        return info
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        info["is_encrypted"]       = bool(reader.is_encrypted)
        info["page_count"]         = len(reader.pages)
        meta = reader.metadata or {}
        info["pdf_metadata_title"]  = str(meta.get("/Title",  "") or "")[:300]
        info["pdf_metadata_author"] = str(meta.get("/Author", "") or "")[:300]
        if reader.pages:
            text = reader.pages[0].extract_text() or ""
            dm = re.search(r"\b10\.\d{4,9}/[-._;()/:A-Z0-9]+\b", text, flags=re.I)
            if dm:
                info["first_page_doi"] = dm.group(0).rstrip(".,;)").lower()
        info["pdf_read_ok"] = True
    except ImportError:
        # pypdf 未安装时退化为只看头部
        info["pdf_read_ok"] = info["pdf_header_valid"]
    except Exception as e:
        info["pdf_error"] = f"{type(e).__name__}: {e}"
    return info


def validate_pdf_file(path: Path, min_kb: int = 20, compute_sha256: bool = True) -> dict:
    """
    单文件完整校验：复用 build_manifest.py 的全部字段。
    包含：头部、大小、页数、加密、首页DOI、PDF元数据标题/作者、文件名解析。
    """
    stat = path.stat()
    parsed   = _parse_filename(path)
    pdf_info = _read_pdf_info(path)
    sha256   = _sha256_file(path) if compute_sha256 else ""
    paper_id = hashlib.sha1(path.name.encode("utf-8")).hexdigest()[:16]

    valid = (
        pdf_info["pdf_header_valid"]
        and stat.st_size > min_kb * 1024
    )
    return {
        "paper_id":          paper_id,
        "file_name":         path.name,
        "file_size_kb":      round(stat.st_size / 1024, 1),
        "file_size_mb":      round(stat.st_size / 1024 / 1024, 3),
        "valid":             valid,
        "sha256":            sha256,
        **parsed,
        **pdf_info,
    }


def scan_pdf_folder(folder: Path, min_kb: int = 20, compute_sha256: bool = True) -> list[dict]:
    """
    扫描目录下全部 PDF，完整复用 build_manifest.py 逻辑：
    文件名解析 + PDF头部 + 页数 + 首页DOI + SHA-256去重 + 加密检测。
    """
    results  = []
    sha_seen: dict[str, str] = {}

    all_pdfs = sorted(folder.rglob("*.pdf"))
    for pdf_path in all_pdfs:
        try:
            r = validate_pdf_file(pdf_path, min_kb, compute_sha256)
        except Exception as e:
            r = {
                "file_name":  pdf_path.name,
                "file_size_kb": 0,
                "valid":      False,
                "pdf_error":  str(e),
            }
        r["relative_path"] = str(pdf_path.relative_to(folder))

        # SHA-256 去重（与 build_manifest.py 一致）
        sha = r.get("sha256", "")
        if sha and sha in sha_seen:
            r["duplicate_of"] = sha_seen[sha]
        else:
            r["duplicate_of"] = ""
            if sha:
                sha_seen[sha] = r["file_name"]

        results.append(r)
    return results


def download_one(doi: str, url: str, publisher: str, out_dir: str, min_kb: int = 20, timeout: int = 60) -> dict:
    try:
        import requests
    except ImportError:
        return {"doi": doi, "status": "error", "reason": "未安装 requests，请运行 pip install requests"}

    out_path = Path(out_dir) / publisher.replace("/", "_") / safe_filename(doi)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # 已存在则跳过
    if out_path.exists() and out_path.stat().st_size > min_kb * 1024:
        v = validate_pdf_bytes(out_path.read_bytes(), min_kb)
        if v["valid"]:
            return {"doi": doi, "status": "skipped_exists", "file": str(out_path),
                    "size_kb": v["size_kb"], "sha256": v["sha256"], "reason": "文件已存在"}

    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
    except Exception as e:
        return {"doi": doi, "status": "error", "reason": str(e)[:120]}

    if resp.status_code in (401, 403):
        return {"doi": doi, "status": "no_access", "file": "", "reason": f"HTTP {resp.status_code} 需机构认证/VPN"}
    if resp.status_code == 404:
        return {"doi": doi, "status": "not_found", "file": "", "reason": "HTTP 404 资源不存在"}
    if resp.status_code != 200:
        return {"doi": doi, "status": "http_error", "file": "", "reason": f"HTTP {resp.status_code}"}

    v = validate_pdf_bytes(resp.content, min_kb)
    if not v["valid"]:
        # 把验证失败的内容也保存下来（存到诊断子目录），方便排查到底下载到了什么
        # 例如付费墙返回的HTML提示页、跳转页等，常见大小几十~几百KB，肉眼一看便知
        debug_dir = Path(out_dir) / "_invalid_debug" / publisher.replace("/", "_")
        debug_dir.mkdir(parents=True, exist_ok=True)
        ctype = resp.headers.get("content-type", "").lower()
        ext = ".html" if "html" in ctype else (".pdf" if "pdf" in ctype else ".bin")
        debug_path = debug_dir / (safe_filename(doi).rsplit(".", 1)[0] + ext)
        try:
            debug_path.write_bytes(resp.content)
        except Exception:
            debug_path = None

        # 关键修复：很多出版商对无权限请求返回 HTTP 200 + HTML登录/付费墙页面，
        # 而不是老老实实返回401/403，导致前面的状态码判断完全失效。
        # 这里对HTML内容做关键词嗅探，尽量把这类"伪装成200"的拒绝访问识别出来，
        # 而不是含糊地报"非PDF内容"让人摸不着头脑。
        status = "invalid_pdf"
        reason = v["reason"] + f"（实际Content-Type: {ctype or '未知'}，已保存供排查）"
        if "html" in ctype:
            try:
                text_lower = resp.content[:20000].decode("utf-8", errors="ignore").lower()
            except Exception:
                text_lower = ""
            paywall_keywords = [
                "sign in", "log in", "purchase", "subscribe", "access denied",
                "institutional access", "get access", "buy this article",
                "您没有权限", "请登录", "购买", "订阅", "access this article",
            ]
            if any(kw in text_lower for kw in paywall_keywords):
                status = "likely_paywall"
                reason = "HTTP 200 但内容是登录/付费墙页面（非真正403），疑似无访问权限"

        return {
            "doi": doi, "status": status, "file": str(debug_path) if debug_path else "",
            "reason": reason, "size_kb": v["size_kb"],
        }

    out_path.write_bytes(resp.content)
    return {"doi": doi, "status": "success", "file": str(out_path),
            "size_kb": v["size_kb"], "sha256": v["sha256"], "reason": ""}


def stat_card(col, value, label, desc=""):
    with col:
        st.markdown(
            f'<div class="stat-card"><div class="stat-label">{label}</div>'
            f'<div class="stat-value">{value}</div>'
            f'<div class="stat-desc">{desc}</div></div>',
            unsafe_allow_html=True,
        )


def check_chrome_debug_port(host: str = "localhost", port: int = 9222) -> bool:
    """检测调试模式Chrome是否在线（webvpn_downloader3.py能否接管它）。"""
    import socket
    try:
        s = socket.create_connection((host, port), timeout=2)
        s.close()
        return True
    except Exception:
        return False


def write_webvpn_queue_csv(df_rows: pd.DataFrame, out_path: Path) -> None:
    """把待下载的WebVPN队列写成 webvpn_downloader3.py 能直接读取的CSV格式。"""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    export_df = pd.DataFrame({
        "doi": df_rows["doi"].astype(str),
        "publisher_group": df_rows.get("出版商", df_rows.get("publisher", "")).astype(str),
        "title": df_rows.get("title", df_rows.get("标题", "")).astype(str) if "title" in df_rows.columns or "标题" in df_rows.columns else "",
    })
    export_df.to_csv(out_path, index=False, encoding="utf-8-sig")


def run_webvpn_downloader(queue_csv: Path, out_dir: str, limit: int, min_kb: int,
                           timeout: int, delay: float, script_path: str) -> subprocess.Popen:
    """以子进程方式启动 webvpn_downloader3.py，实时把stdout流式吐给调用方。"""
    cmd = [
        sys.executable, script_path,
        "--queue", str(queue_csv),
        "--out-dir", out_dir,
        "--limit", str(limit),
        "--min-kb", str(min_kb),
        "--timeout", str(timeout),
        "--delay", str(delay),
        "--debug",
    ]
    return subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        bufsize=1,
    )


def parse_webvpn_log(log_path: Path) -> pd.DataFrame:
    """读取 webvpn_downloader3.py 生成的日志CSV。"""
    if not log_path.exists():
        return pd.DataFrame()
    try:
        return pd.read_csv(log_path, encoding="utf-8-sig")
    except Exception:
        return pd.DataFrame()


def read_log_tail(path: str, n: int = 50) -> pd.DataFrame:
    """读取CSV文件最后n行（自动抠取数据页面用于展示manifest等结果）。"""
    try:
        p = Path(path)
        if not p.exists():
            return pd.DataFrame()
        return pd.read_csv(p, encoding="utf-8-sig").tail(n)
    except Exception:
        return pd.DataFrame()


def run_cmd_sync(cmd: list, cwd: str = None, timeout: int = 300) -> tuple:
    """同步执行命令，阻塞直到完成，返回(退出码, 完整输出)。"""
    try:
        r = subprocess.run(
            cmd, cwd=cwd, capture_output=True,
            text=True, encoding="utf-8", errors="replace",
            timeout=timeout, shell=False,
        )
        return r.returncode, (r.stdout + r.stderr)
    except subprocess.TimeoutExpired:
        return -1, f"超时（>{timeout}s）"
    except Exception as e:
        return -1, str(e)


def run_cmd_bg(cmd: list, cwd: str = None) -> subprocess.Popen:
    """后台启动命令，返回Popen对象供调用方流式读取stdout。"""
    return subprocess.Popen(
        cmd, cwd=cwd,
        stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
        text=True, encoding="utf-8", errors="replace", shell=False,
    )


# ══════════════════════════════════════════════════
#  数据可视化分析：数值解析与单位归一化工具
# ══════════════════════════════════════════════════

# ══════════════════════════════════════════════════
#  数值与单位解析
# ══════════════════════════════════════════════════

def parse_numeric_with_unit(raw: str) -> dict:
    """
    把一个可能带单位/带百分号/带范围的原始字符串，解析成结构化结果。
    返回 dict: {
        "raw": 原始文本,
        "kind": "single" | "range" | "unparseable",
        "value": 单点数值 (kind=single时有效),
        "value_min": 区间下界 (kind=range时有效),
        "value_max": 区间上界 (kind=range时有效),
        "unit_raw": 解析出的原始单位文本,
    }
    """
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return {"raw": raw, "kind": "unparseable", "value": None,
                "value_min": None, "value_max": None, "unit_raw": None}

    text = str(raw).strip()
    if not text:
        return {"raw": raw, "kind": "unparseable", "value": None,
                "value_min": None, "value_max": None, "unit_raw": None}

    # 区间形式: "90-204 h", "room temperature to 950 °C"
    range_match = re.search(
        r"(?:room temperature|RT|(-?\d+\.?\d*))\s*(?:to|-|–|~)\s*(-?\d+\.?\d*)\s*([%a-zA-Z°/().\s]*)$",
        text, re.IGNORECASE,
    )
    if range_match:
        low_str, high_str, unit = range_match.group(1), range_match.group(2), range_match.group(3)
        low = 25.0 if low_str is None else float(low_str)  # "room temperature" 近似为25°C
        try:
            high = float(high_str)
            return {"raw": raw, "kind": "range", "value": None,
                    "value_min": low, "value_max": high, "unit_raw": unit.strip() or None}
        except ValueError:
            pass

    # 单点数值: "523 K", "280°C", "55%", "66.0", "64"
    single_match = re.search(r"(-?\d+\.?\d*)\s*([%a-zA-Z°/().]*)", text)
    if single_match:
        try:
            value = float(single_match.group(1))
            unit = single_match.group(2).strip() or None
            return {"raw": raw, "kind": "single", "value": value,
                     "value_min": None, "value_max": None, "unit_raw": unit}
        except ValueError:
            pass

    # 完全无法解析（如 "tenfold increase", "TOS (h)" 这种字段名误填）
    return {"raw": raw, "kind": "unparseable", "value": None,
            "value_min": None, "value_max": None, "unit_raw": None}


# ══════════════════════════════════════════════════
#  单位归一化（自动识别常见单位并统一换算）
# ══════════════════════════════════════════════════

def _safe_unit_str(unit_raw) -> str:
    """统一把None/NaN/空值转成空字符串，避免下游 .lower() 报错。"""
    if unit_raw is None:
        return ""
    if isinstance(unit_raw, float) and pd.isna(unit_raw):
        return ""
    return str(unit_raw)


def normalize_temperature(value: float, unit_raw) -> Optional[float]:
    """统一换算成 °C。"""
    if value is None:
        return None
    u = _safe_unit_str(unit_raw).lower().replace(" ", "")
    if "k" == u or (u.endswith("k") and "c" not in u):  # Kelvin
        return round(value - 273.15, 2)
    if "f" in u:  # Fahrenheit（少见，但兜底）
        return round((value - 32) * 5 / 9, 2)
    # 默认/已是 °C 或无单位时按摄氏度处理
    return value


def normalize_pressure(value: float, unit_raw) -> Optional[float]:
    """统一换算成 MPa。"""
    if value is None:
        return None
    u = _safe_unit_str(unit_raw).lower().replace(" ", "")
    if "bar" in u:
        return round(value * 0.1, 4)
    if "atm" in u:
        return round(value * 0.101325, 4)
    if "psi" in u:
        return round(value * 0.00689476, 4)
    if "kpa" in u:
        return round(value / 1000, 4)
    # 默认/已是 MPa
    return value


def normalize_time_hours(value: float, unit_raw) -> Optional[float]:
    """统一换算成小时。"""
    if value is None:
        return None
    u = _safe_unit_str(unit_raw).lower().replace(" ", "")
    if "min" in u:
        return round(value / 60, 3)
    if "day" in u or u == "d":
        return round(value * 24, 2)
    # 默认/已是小时（h）
    return value


UNIT_NORMALIZERS = {
    "temperature": (normalize_temperature, "°C"),
    "pressure": (normalize_pressure, "MPa"),
    "time_on_stream": (normalize_time_hours, "h"),
}


def parse_and_normalize_column(series: pd.Series, field_name: str) -> pd.DataFrame:
    """
    对一整列原始文本（如 reaction_conditions 的 temperature 列）做解析+归一化。
    返回一个DataFrame，包含: raw, kind, value_norm, value_min_norm, value_max_norm, target_unit
    无法解析的行 value_norm 等均为 None，kind='unparseable'，原文保留在 raw 列供查看。
    """
    parsed = series.apply(parse_numeric_with_unit).apply(pd.Series)

    normalizer_info = UNIT_NORMALIZERS.get(field_name)
    if normalizer_info is None:
        # 没有定义归一化规则的字段（如 h2_co_ratio 这种本身已是无量纲比值），原样透传数值
        parsed["value_norm"] = parsed["value"]
        parsed["value_min_norm"] = parsed["value_min"]
        parsed["value_max_norm"] = parsed["value_max"]
        parsed["target_unit"] = parsed["unit_raw"]
        return parsed

    normalizer_fn, target_unit = normalizer_info
    parsed["value_norm"] = parsed.apply(
        lambda r: normalizer_fn(r["value"], r["unit_raw"]) if r["kind"] == "single" else None, axis=1
    )
    parsed["value_min_norm"] = parsed.apply(
        lambda r: normalizer_fn(r["value_min"], r["unit_raw"]) if r["kind"] == "range" else None, axis=1
    )
    parsed["value_max_norm"] = parsed.apply(
        lambda r: normalizer_fn(r["value_max"], r["unit_raw"]) if r["kind"] == "range" else None, axis=1
    )
    parsed["target_unit"] = target_unit
    return parsed


# ══════════════════════════════════════════════════
#  长表 → 各阶段独立宽表
# ══════════════════════════════════════════════════

def split_stages(df_long: pd.DataFrame) -> dict:
    """把长表按stage拆分成4个独立的宽表（去掉全空列），并对数值字段做单位归一化。"""
    stages = {}

    # metadata
    meta = df_long[df_long["stage"] == "metadata"].dropna(axis=1, how="all").copy()
    stages["metadata"] = meta

    # catalyst
    cat = df_long[df_long["stage"] == "catalyst"].dropna(axis=1, how="all").copy()
    stages["catalyst"] = cat

    # reaction_conditions —— 对temperature/pressure/time_on_stream做单位归一化
    cond = df_long[df_long["stage"] == "reaction_conditions"].dropna(axis=1, how="all").copy()
    for field in ["temperature", "pressure", "time_on_stream"]:
        if field in cond.columns:
            parsed = parse_and_normalize_column(cond[field], field)
            cond[f"{field}_value"] = parsed["value_norm"]
            cond[f"{field}_min"] = parsed["value_min_norm"]
            cond[f"{field}_max"] = parsed["value_max_norm"]
            cond[f"{field}_kind"] = parsed["kind"]
            cond[f"{field}_unit"] = parsed["target_unit"]
    stages["reaction_conditions"] = cond

    # performance —— 对value做解析（百分比/数值/区间/无法解析）
    perf = df_long[df_long["stage"] == "performance"].dropna(axis=1, how="all").copy()
    if "value" in perf.columns:
        parsed = parse_and_normalize_column(perf["value"], "value")  # 无归一化规则，原样透传
        perf["value_parsed"] = parsed["value_norm"]
        perf["value_kind"] = parsed["kind"]
        perf["value_min_parsed"] = parsed["value_min_norm"]
        perf["value_max_parsed"] = parsed["value_max_norm"]
    stages["performance"] = perf

    return stages


def build_plot_ready_performance(perf_df: pd.DataFrame, paper_meta: pd.DataFrame) -> pd.DataFrame:
    """
    把performance宽表里 kind=='single' 的行筛出来，关联上论文年份/期刊，
    生成可以直接画图的"整洁"DataFrame：每行一个可比较的数据点。

    metric 处理：论文原文写法五花八门（如"CO Initial Conversion"、"CO Steady-State
    Conversion"都属于conversion大类），全部保留会导致数据量大了之后选项爆炸、无法横向
    汇总比较。这里做两件事：
      - metric_raw：保留论文原文，供查证/悬停提示用，不丢失任何信息
      - metric：归并到 conversion/selectivity/yield/STY/productivity/其他 六个大类，
                图表的分组、下拉框默认用这一列，确保数据规模变大后仍然可比较
    """
    plot_df = perf_df[perf_df.get("value_kind") == "single"].copy()
    if plot_df.empty:
        return plot_df

    plot_df = plot_df.rename(columns={"value_parsed": "value_numeric"})
    if "metric" in plot_df.columns:
        plot_df["metric_raw"] = plot_df["metric"]
        plot_df["metric"] = plot_df["metric"].apply(normalize_metric_category)

    keep_cols = ["paper_id", "performance_id", "catalyst_ref", "metric", "metric_raw",
                 "species_or_range", "value_numeric", "unit", "confidence"]
    keep_cols = [c for c in keep_cols if c in plot_df.columns]
    plot_df = plot_df[keep_cols]

    if not paper_meta.empty and "paper_id" in paper_meta.columns:
        meta_cols = [c for c in ["paper_id", "metadata_year", "metadata_journal", "metadata_doi"] if c in paper_meta.columns]
        plot_df = plot_df.merge(paper_meta[meta_cols], on="paper_id", how="left")
        if "metadata_doi" in plot_df.columns:
            plot_df = plot_df.rename(columns={"metadata_doi": "doi"})

    return plot_df


# 指标大类归并规则：把论文原文的开放词汇映射回 04_performance.md 定义的5大类。
# 按关键词匹配，顺序很重要——更具体的关键词写在前面，避免被泛化关键词提前命中。
METRIC_CATEGORY_RULES = [
    ("STY", ["sty", "space time yield", "space-time yield"]),
    ("productivity", ["productivity"]),
    ("yield", ["yield"]),
    ("selectivity", ["selectivity"]),
    ("conversion", ["conversion"]),
]
METRIC_CATEGORY_OTHER = "其他"


def normalize_metric_category(raw_metric) -> str:
    """
    把单个metric原文字符串，归并到5大类之一，或"其他"。
    例如 "CO Initial Conversion" -> "conversion"
         "C5+ Selectivity" / "C5+ selectivity" -> "selectivity" （大小写不敏感）
         "FT/(FT + WGS)" -> "其他" （不属于5大类任何一种）
    """
    if raw_metric is None or (isinstance(raw_metric, float) and pd.isna(raw_metric)):
        return METRIC_CATEGORY_OTHER
    text_lower = str(raw_metric).lower()
    for category, keywords in METRIC_CATEGORY_RULES:
        if any(kw in text_lower for kw in keywords):
            return category
    return METRIC_CATEGORY_OTHER


def get_unparseable_rows(stages: dict) -> pd.DataFrame:
    """收集所有阶段里"无法解析为数值"的原始行，供页面单独列表展示。"""
    rows = []

    perf = stages.get("performance", pd.DataFrame())
    if not perf.empty and "value_kind" in perf.columns:
        bad = perf[perf["value_kind"] == "unparseable"]
        for _, r in bad.iterrows():
            rows.append({
                "paper_id": r.get("paper_id"), "字段": "performance.value",
                "原始文本": r.get("value"), "metric": r.get("metric"),
                "species_or_range": r.get("species_or_range"),
            })

    cond = stages.get("reaction_conditions", pd.DataFrame())
    for field in ["temperature", "pressure", "time_on_stream"]:
        kind_col = f"{field}_kind"
        if not cond.empty and kind_col in cond.columns:
            bad = cond[cond[kind_col] == "unparseable"]
            for _, r in bad.iterrows():
                rows.append({
                    "paper_id": r.get("paper_id"), "字段": f"reaction_conditions.{field}",
                    "原始文本": r.get(field), "metric": None, "species_or_range": None,
                })

    return pd.DataFrame(rows)


# 反应条件字段 → 展示名 + 单位，用于图表下拉选项和坐标轴标签
CONDITION_FIELD_LABELS = {
    "temperature": ("温度", "°C"),
    "pressure": ("压力", "MPa"),
    "time_on_stream": ("运行时长", "h"),
}


def build_plot_ready_conditions(cond_df: pd.DataFrame, paper_meta: pd.DataFrame) -> pd.DataFrame:
    """
    把reaction_conditions宽表里的temperature/pressure/time_on_stream三个字段，
    转成和performance同样的"长格式"：每行一个(paper_id, 字段名, 数值)三元组，
    这样可以复用同一套箱线图/折线图代码。
    只保留 kind=='single' 的行（区间值如"90-204h"不在此处展开，避免歧义）。
    """
    if cond_df.empty:
        return pd.DataFrame()

    rows = []
    for field, (label, unit) in CONDITION_FIELD_LABELS.items():
        value_col = f"{field}_value"
        kind_col = f"{field}_kind"
        if value_col not in cond_df.columns:
            continue
        sub = cond_df[cond_df.get(kind_col) == "single"]
        for _, r in sub.iterrows():
            rows.append({
                "paper_id": r.get("paper_id"),
                "condition_id": r.get("condition_id"),
                "catalyst_ref": r.get("catalyst_ref"),
                "metric": label,            # 复用"metric"这个列名，方便和performance共用画图代码
                "species_or_range": None,
                "value_numeric": r.get(value_col),
                "unit": unit,
                "confidence": None,
            })

    plot_df = pd.DataFrame(rows)
    if plot_df.empty:
        return plot_df

    if not paper_meta.empty and "paper_id" in paper_meta.columns:
        meta_cols = [c for c in ["paper_id", "metadata_year", "metadata_journal", "metadata_doi"] if c in paper_meta.columns]
        plot_df = plot_df.merge(paper_meta[meta_cols], on="paper_id", how="left")
        if "metadata_doi" in plot_df.columns:
            plot_df = plot_df.rename(columns={"metadata_doi": "doi"})

    return plot_df


def join_conditions_with_performance(cond_plot_df: pd.DataFrame, perf_plot_df: pd.DataFrame) -> pd.DataFrame:
    """
    按 paper_id + catalyst_ref 关联反应条件和性能数据，
    用于"温度 vs 转化率"这类跨阶段散点图。
    """
    if cond_plot_df.empty or perf_plot_df.empty:
        return pd.DataFrame()

    cond_wide = cond_plot_df.pivot_table(
        index=["paper_id", "catalyst_ref"], columns="metric", values="value_numeric", aggfunc="first"
    ).reset_index()

    merge_keys = ["paper_id", "catalyst_ref"]
    if "catalyst_ref" not in perf_plot_df.columns or "catalyst_ref" not in cond_wide.columns:
        return pd.DataFrame()

    merged = perf_plot_df.merge(cond_wide, on=merge_keys, how="inner")
    return merged




def plotly_chart_with_doi(fig, source_df: pd.DataFrame, chart_key: str,
                          doi_col: str = "doi", paper_id_col: str = "paper_id"):
    """
    通用的"带DOI溯源"图表渲染函数，适用于箱线图/折线图/散点图/云雨图所有图表类型。
    点击图中任意数据点后，在图表下方显示该论文的DOI跳转按钮。

    source_df：与图表数据对应的源DataFrame，需含 paper_id 列（以及 doi 列，如有）。
    原理：on_select 捕获点击事件 → 用 point_index 在 source_df 里定位 → 读取DOI → 显示跳转按钮。
    """
    event = st.plotly_chart(fig, use_container_width=True, on_select="rerun", key=chart_key)

    if event and event.selection and event.selection.get("points"):
        pt = event.selection["points"][0]
        pt_idx = pt.get("point_index", -1)

        # 尝试从source_df里找到对应行
        row = None
        if 0 <= pt_idx < len(source_df):
            row = source_df.iloc[pt_idx]

        if row is not None:
            doi_raw = ""
            if doi_col in source_df.columns:
                doi_raw = str(row.get(doi_col, "") or "").strip()
                doi_raw = doi_raw.replace("nan","").replace("None","")

            paper_id = str(row.get(paper_id_col, "")) if paper_id_col in source_df.columns else ""

            if doi_raw:
                doi_url = doi_raw if doi_raw.startswith("http") else f"https://doi.org/{doi_raw}"
                cols_doi = st.columns([3, 1])
                with cols_doi[0]:
                    st.info(f"📄 已选中 | paper_id: `{paper_id}` | DOI: `{doi_raw}`")
                with cols_doi[1]:
                    st.link_button("🔗 打开论文", doi_url, use_container_width=True)
            else:
                st.caption(f"已选中 paper_id: `{paper_id}`（暂无DOI信息）")
    else:
        st.caption("💡 点击图中任意数据点可获得该论文的 DOI 跳转链接")


# 保留旧名称作为别名，避免漏改的调用处报错
def scatter_with_doi_links(fig, plot_data: pd.DataFrame, doi_col: str = "doi",
                           height: int = 480, chart_key: str = "doi_scatter"):
    fig.update_layout(height=height)
    plotly_chart_with_doi(fig, plot_data, chart_key=chart_key, doi_col=doi_col)


def build_plot_ready_metadata(meta_df: pd.DataFrame) -> pd.DataFrame:
    """
    把metadata阶段数据整理成可画图的格式。
    主要用途：按年份统计论文数量、各route_type/document_type的分布。
    返回每篇论文一行，包含year、journal、route_type、document_type等字段。
    """
    if meta_df.empty:
        return pd.DataFrame()

    cols_wanted = {
        "paper_id": "paper_id",
        "metadata_year": "year",
        "metadata_journal": "journal",
        "metadata_document_type": "document_type",
        "metadata_route_type": "route_type",
        "metadata_doi": "doi",
    }
    available = {k: v for k, v in cols_wanted.items() if k in meta_df.columns}
    plot_df = meta_df[list(available.keys())].rename(columns=available).copy()

    # year转数值
    if "year" in plot_df.columns:
        plot_df["year"] = pd.to_numeric(plot_df["year"], errors="coerce")

    # route_type字段是JSON列表字符串，提取第一个value值
    if "route_type" in plot_df.columns:
        def extract_first_route(raw):
            if pd.isna(raw) or not str(raw).strip():
                return None
            import re as _re
            m = _re.search(r"'value':\s*'([^']+)'", str(raw))
            return m.group(1) if m else str(raw)[:50]
        plot_df["route_type"] = plot_df["route_type"].apply(extract_first_route)

    return plot_df.dropna(subset=["paper_id"])


def build_plot_ready_catalyst(cat_df: pd.DataFrame, meta_df: pd.DataFrame) -> pd.DataFrame:
    """
    把catalyst阶段数据整理成可画图的格式。
    把active_metal/support/promoter/preparation_method转成统一的长格式
    （metric=字段名, value_text=字段值），方便用箱线图/条形图看各类材料的分布。
    """
    if cat_df.empty:
        return pd.DataFrame()

    cat_fields = {
        "active_metal":       "活性金属",
        "support":            "载体",
        "promoter":           "助剂",
        "preparation_method": "制备方法",
    }

    rows = []
    for field, label in cat_fields.items():
        if field not in cat_df.columns:
            continue
        sub = cat_df[["paper_id", field]].dropna(subset=[field])
        for _, r in sub.iterrows():
            val = str(r[field]).strip()
            if val and val.lower() not in ("nan", "none", "null", ""):
                rows.append({
                    "paper_id": r["paper_id"],
                    "metric": label,
                    "value_text": val,
                })

    if not rows:
        return pd.DataFrame()

    plot_df = pd.DataFrame(rows)

    # 关联年份和DOI信息
    if not meta_df.empty and "paper_id" in meta_df.columns:
        year_col = "metadata_year" if "metadata_year" in meta_df.columns else None
        if year_col:
            year_map = meta_df.set_index("paper_id")[year_col].apply(
                lambda x: pd.to_numeric(x, errors="coerce")
            )
            plot_df["metadata_year"] = plot_df["paper_id"].map(year_map)
        # 关联DOI
        doi_col = "metadata_doi" if "metadata_doi" in meta_df.columns else None
        if doi_col:
            doi_map = meta_df.set_index("paper_id")[doi_col].astype(str)
            plot_df["doi"] = plot_df["paper_id"].map(doi_map)

    return plot_df


def pipeline_bar(current: int, steps: Optional[list] = None):
    """显示流程进度条，current=1~N。不传steps则用默认的下载4步流程。"""
    if steps is None:
        steps = ["① DOI 分类", "② 自动下载", "③ PDF 校验", "④ 导出结果"]
    parts = []
    for i, s in enumerate(steps, 1):
        if i < current:
            cls = "done"
        elif i == current:
            cls = "active"
        else:
            cls = "pending"
        parts.append(f'<div class="pip-step {cls}">{s}</div>')
        if i < len(steps):
            parts.append('<div class="pip-arrow">›</div>')
    st.markdown(
        f'<div class="pipeline-bar">{"".join(parts)}</div>',
        unsafe_allow_html=True,
    )


# ══════════════════════════════════════════════════════════════════
#  顶部横排导航栏（替代侧边栏）
# ══════════════════════════════════════════════════════════════════

_PAGES = ["📊 平台总览", "🗃️ 数据库浏览", "📥 论文自动下载", "🔬 自动抠取数据", "📈 数据可视化分析", "📖 使用说明"]

if "page" not in st.session_state:
    st.session_state["page"] = _PAGES[0]

# ── Logo + 标题区 ──
import base64 as _b64
_img_tag = ""
# 依次尝试常见文件名，兼容本地和云端部署
_logo_candidates = ["logo.jpg", "logo.png", "logo.jpeg",
                    "微信图片_20260623160903.jpg", "微信图片_20260623160903.png"]
for _logo_name in _logo_candidates:
    _logo_path = Path(__file__).parent / _logo_name
    if _logo_path.exists():
        try:
            with open(_logo_path, "rb") as _f:
                _img_b64 = _b64.b64encode(_f.read()).decode()
            _ext = _logo_path.suffix.lower().lstrip(".")
            _mime = "jpeg" if _ext in ("jpg", "jpeg") else _ext
            _img_tag = f'<img src="data:image/{_mime};base64,{_img_b64}" style="height:52px;width:auto;border-radius:6px;flex-shrink:0;" />'
        except Exception:
            pass
        break

st.markdown(f"""
<div style="background:linear-gradient(135deg,#0D2B5E 0%,#1565C0 100%);
            padding:14px 32px 0 24px;
            box-shadow:0 2px 8px rgba(0,0,0,0.25);
            margin-bottom:0;">
  <div style="display:flex;align-items:center;gap:16px;padding-bottom:10px;">
    {_img_tag}
    <div>
      <div style="color:#fff;font-size:1.35rem;font-weight:700;line-height:1.25;letter-spacing:-0.01em;">
        费托合成催化剂材料数据自动挖掘与大数据分析平台
      </div>
      <div style="color:rgba(255,255,255,0.75);font-size:0.78rem;margin-top:2px;">
        实现文献的自动批量下载、文献内数据的自动摘取以及数据可视化分析。
      </div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── 导航标签行：用原生按钮 + CSS 完全伪装成导航栏样式 ──
st.markdown("""
<style>
/* 导航栏按钮容器背景 */
div[data-testid="stHorizontalBlock"].navbar-row > div[data-testid="stColumn"] {
    padding: 0 !important;
}
/* 把所有导航按钮统一样式 */
.navbar-row button {
    background: linear-gradient(135deg,#0D2B5E 0%,#1565C0 100%) !important;
    color: rgba(255,255,255,0.72) !important;
    border: none !important;
    border-radius: 0 !important;
    border-bottom: 3px solid transparent !important;
    font-size: 0.95rem !important;
    font-weight: 500 !important;
    padding: 10px 4px !important;
    width: 100% !important;
    transition: all 0.15s !important;
}
.navbar-row button:hover {
    color: #fff !important;
    background: linear-gradient(135deg,#0D2B5E 0%,#1976D2 100%) !important;
    border-bottom: 3px solid rgba(255,255,255,0.5) !important;
}
/* 当前激活的按钮 */
.navbar-row button.active-nav,
.navbar-row button[data-active="true"] {
    color: #fff !important;
    border-bottom: 3px solid #fff !important;
    font-weight: 700 !important;
}
/* 去掉按钮外层多余间距 */
.navbar-row .stButton { margin: 0 !important; }
.navbar-row { gap: 0 !important; background: linear-gradient(135deg,#0D2B5E 0%,#1565C0 100%); margin-bottom: 0 !important; }
</style>
""", unsafe_allow_html=True)

# 渲染导航按钮
_nav_cols = st.columns(len(_PAGES))
for _i, (_col, _p) in enumerate(zip(_nav_cols, _PAGES)):
    with _col:
        if st.button(_p, key=f"nav_btn_{_i}", use_container_width=True):
            st.session_state["page"] = _p
            st.rerun()

# 用 JS 给当前激活按钮加样式
_cur = st.session_state["page"]
_active_idx = _PAGES.index(_cur)
st.markdown(f"""
<script>
(function() {{
    function styleNavBtns() {{
        var btns = document.querySelectorAll('.navbar-row button');
        if (btns.length < {len(_PAGES)}) {{ setTimeout(styleNavBtns, 80); return; }}
        btns.forEach(function(b, i) {{
            if (i === {_active_idx}) {{
                b.style.color = '#fff';
                b.style.borderBottom = '3px solid #fff';
                b.style.fontWeight = '700';
            }} else {{
                b.style.color = 'rgba(255,255,255,0.72)';
                b.style.borderBottom = '3px solid transparent';
                b.style.fontWeight = '500';
            }}
        }});
    }}
    styleNavBtns();
}})();
</script>
""", unsafe_allow_html=True)

# 给按钮行容器打上 class（Streamlit 不支持直接设 class，用 JS 补）
st.markdown("""
<script>
(function() {
    function tagNavRow() {
        var allRows = document.querySelectorAll('[data-testid="stHorizontalBlock"]');
        allRows.forEach(function(row) {
            var btns = row.querySelectorAll('button');
            if (btns.length === 6) { row.classList.add('navbar-row'); }
        });
        if (!document.querySelector('.navbar-row')) { setTimeout(tagNavRow, 80); }
    }
    tagNavRow();
})();
</script>
""", unsafe_allow_html=True)

page = st.session_state["page"]

# ══════════════════════════════════════════════════════════════════
#  Hero Banner
# ══════════════════════════════════════════════════════════════════
HERO_SUBS = {
    "📊 平台总览":       "实现文献的自动批量下载、文献内数据的自动摘取以及数据可视化分析。",
    "🗃️ 数据库浏览":     "按催化剂家族、活性金属、原料气类型、年份和反应条件筛选，导出 CSV。",
    "📥 论文自动下载":   "DOI 分类 → 自动下载 → PDF 校验 → 导出结果，全流程一站式完成。",
    "🔬 自动抠取数据":   "对已下载 PDF 执行多阶段结构化抽取，输出催化剂、反应条件与性能数据。",
    "📈 数据可视化分析": "对抽取结果进行多维度图表分析。",
    "📖 使用说明":       "说明页面各模块含义、数据文件位置和生产应用边界。",
}
st.markdown(f"""
<div class="hero-banner">
  <div class="hero-tag">费托合成催化剂材料数据平台</div>
  <div class="hero-title">费托合成催化剂材料数据自动挖掘与大数据分析平台</div>
  <div class="hero-sub">{HERO_SUBS.get(page, "")}</div>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
#  页面：平台总览
# ══════════════════════════════════════════════════════════════════
if page == "📊 平台总览":

    cols = st.columns(5)
    for col, val, lbl, desc in zip(cols,
        ["1,627", "1,627", "2004–2026", "502 / 1,125", "11"],
        ["文献总量", "有效 PDF", "年份范围", "OA / 非OA", "出版商数"],
        ["PDF 全量已扫描", "全部有效 0读取错误", "主力年份2014-2025", "OA 30.9%·非OA 69.1%", "ACS·Wiley·Springer等"],
    ):
        stat_card(col, val, lbl, desc)

    st.markdown("")
    left, right = st.columns([3, 2])
    with left:
        st.markdown('<div class="section-title">出版商分布</div>', unsafe_allow_html=True)
        st.bar_chart(pd.DataFrame({
            "出版商": ["ACS","Wiley","Springer Nature","RSC","Elsevier","其他"],
            "文献数": [520, 391, 390, 220, 67, 39],
        }).set_index("出版商"), color="#1565C0", height=260)

    with right:
        st.markdown('<div class="section-title">论文自动下载：效率对比</div>', unsafe_allow_html=True)
        TOTAL_PAPERS = 1627
        MANUAL_PER_DAY = 100
        manual_days = round(TOTAL_PAPERS / MANUAL_PER_DAY, 1)
        auto_days = 1

        compare_df = pd.DataFrame({
            "方式": ["自动批量下载AI", "人工逐篇下载"],
            "耗时(天)": [auto_days, manual_days],
        })
        fig_compare = px.bar(
            compare_df, x="方式", y="耗时(天)",
            color_discrete_sequence=["#1565C0"],
            category_orders={"方式": ["自动批量下载AI", "人工逐篇下载"]},
        )
        fig_compare.update_layout(
            height=220, margin=dict(t=10, b=10, l=10, r=10),
            showlegend=False, yaxis_title="耗时（天）", xaxis_title="",
        )
        st.plotly_chart(fig_compare, use_container_width=True)
        st.markdown(
            f'<div style="display:flex;justify-content:space-around;margin-top:6px">'
            f'<div style="text-align:center"><div style="font-size:1.6rem;font-weight:700;color:#2E7D32">{auto_days} 天</div>'
            f'<div style="font-size:0.78rem;color:#666">自动批量下载 {TOTAL_PAPERS} 篇</div>'
            f'<div style="font-size:0.73rem;color:#AAA">效率提升约 {round(manual_days/auto_days, 1)} 倍</div></div>'
            f'<div style="text-align:center"><div style="font-size:1.6rem;font-weight:700;color:#1565C0">{manual_days} 天</div>'
            f'<div style="font-size:0.78rem;color:#666">人工下载（按每天{MANUAL_PER_DAY}篇估算）</div></div>'
            f'</div>',
            unsafe_allow_html=True,
        )



# ══════════════════════════════════════════════════════════════════
#  页面：数据库浏览
# ══════════════════════════════════════════════════════════════════
elif page == "🗃️ 数据库浏览":

    # 优先使用相对路径（云端部署），找不到时回退到本地路径（本地运行）
    _base = Path(__file__).parent
    _db_candidates = [
        _base / "data" / "FT_SAF_catalyst_extraction_wide_table.xlsx",
        _base / "data" / "FT_SAF_catalyst_extraction_wide_table - 副本.csv",
        _base / "data" / "FT_SAF_catalyst_extraction_wide_table.csv",
        _base / "FT_SAF_catalyst_extraction_wide_table.xlsx",
        _base / "FT_SAF_catalyst_extraction_wide_table.csv",
        Path(r"D:\agent\saf_extraction\outputs\FT_SAF_catalyst_extraction_wide_table.xlsx"),
    ]
    _db_found = next((p for p in _db_candidates if p.exists()), None)
    DB_PATH = str(_db_found) if _db_found else str(_base / "data" / "FT_SAF_catalyst_extraction_wide_table.xlsx")

    @st.cache_data(show_spinner="加载数据库…")
    def load_db(path):
        try:
            if not Path(path).exists():
                st.warning(f"⚠️ 数据文件不存在：{path}\n\n请将数据文件放在项目的 `data/` 目录下后重新部署。")
                return pd.DataFrame()
            if path.endswith(".csv"):
                df = pd.read_csv(path, encoding="utf-8-sig")
            else:
                df = pd.read_excel(path)
        except Exception as e:
            st.error(f"读取数据库失败：{e}")
            return pd.DataFrame()
        # 清洗数值列中的明显异常值
        for col in ["reaction_temperature_C", "reaction_pressure_bar",
                    "CO_conversion_pct", "CO2_conversion_pct", "BET_surface_area_m2_g",
                    "metal_particle_size_nm", "H2_CO_ratio"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        if "year" in df.columns:
            df["year"] = pd.to_numeric(df["year"], errors="coerce")
        # 过滤明显错误的异常值
        if "reaction_temperature_C" in df.columns:
            df.loc[df["reaction_temperature_C"] > 1200, "reaction_temperature_C"] = float("nan")
        if "reaction_pressure_bar" in df.columns:
            df.loc[df["reaction_pressure_bar"] > 1000, "reaction_pressure_bar"] = float("nan")
        if "CO_conversion_pct" in df.columns:
            df.loc[df["CO_conversion_pct"] > 100, "CO_conversion_pct"] = float("nan")
        return df

    db = load_db(DB_PATH)

    if db.empty:
        st.stop()

    # ── 提取催化剂家族主类（取第一段）──
    def primary_family(val):
        if pd.isna(val):
            return "未知"
        return str(val).split(";")[0].strip()

    def primary_metal(val):
        if pd.isna(val):
            return "未知"
        return str(val).split(";")[0].strip()

    # 兼容列名不存在的情况
    db["_family_primary"] = db["catalyst_family"].apply(primary_family) if "catalyst_family" in db.columns else "未知"
    db["_metal_primary"]  = db["active_metal"].apply(primary_metal) if "active_metal" in db.columns else "未知"

    # ── 侧边筛选区（左侧列） ──
    left_col, right_col = st.columns([1, 3])

    with left_col:
        st.markdown('<div class="section-title">筛选条件</div>', unsafe_allow_html=True)

        # 年份范围
        yr_min = int(db["year"].dropna().min())
        yr_max = int(db["year"].dropna().max())
        yr_range = st.slider("年份范围", yr_min, yr_max, (2010, yr_max), key="db_yr")

        # 催化剂家族
        family_opts = sorted(db["_family_primary"].dropna().unique())
        sel_family = st.multiselect("催化剂家族", family_opts, key="db_family",
                                    placeholder="全部（不选=全部）")

        # 活性金属
        metal_opts = sorted(db["_metal_primary"].dropna().unique())
        sel_metal = st.multiselect("活性金属", metal_opts, key="db_metal",
                                   placeholder="全部（不选=全部）")

        # 原料气类型（固定显示，不再提供 CO/CO₂ 路由选择）
        st.markdown("**原料气类型**  \n原料气")

        # 反应温度
        t_data = db["reaction_temperature_C"].dropna()
        if len(t_data):
            t_min, t_max = int(t_data.min()), int(t_data.max())
            temp_range = st.slider("反应温度 °C", t_min, t_max, (t_min, t_max), key="db_temp")
        else:
            temp_range = None

        # 关键词搜索
        kw = st.text_input("关键词搜索", placeholder="催化剂名称、载体、论文题目…", key="db_kw")

    # ── 筛选逻辑 ──
    filtered = db.copy()
    filtered = filtered[
        filtered["year"].isna() |
        ((filtered["year"] >= yr_range[0]) & (filtered["year"] <= yr_range[1]))
    ]
    if sel_family:
        filtered = filtered[filtered["_family_primary"].isin(sel_family)]
    if sel_metal:
        filtered = filtered[filtered["_metal_primary"].isin(sel_metal)]
    if temp_range:
        t_mask = (
            filtered["reaction_temperature_C"].isna() |
            ((filtered["reaction_temperature_C"] >= temp_range[0]) &
             (filtered["reaction_temperature_C"] <= temp_range[1]))
        )
        filtered = filtered[t_mask]
    if kw.strip():
        kw_lower = kw.strip().lower()
        search_cols = ["catalyst_name_normalized", "catalyst_name_raw",
                       "support", "paper_title", "catalyst_family"]
        mask = pd.Series(False, index=filtered.index)
        for sc in search_cols:
            if sc in filtered.columns:
                mask |= filtered[sc].astype(str).str.lower().str.contains(kw_lower, na=False)
        filtered = filtered[mask]

    # ── 右侧：统计卡片 + 数据表 + 散点图 ──
    with right_col:

        # 统计卡片行
        n = len(filtered)
        avg_co  = filtered["CO_conversion_pct"].mean()
        avg_co2 = filtered["CO2_conversion_pct"].mean()
        avg_h2co = filtered["H2_CO_ratio"].mean()

        c1, c2, c3, c4 = st.columns(4)
        total_cells = n * 254
        stat_card(c1, f"{total_cells:,}", "筛选后数据量", f"共 {n:,} 条记录 × 254 列")
        stat_card(c2, f"{avg_co:.2f}"  if not pd.isna(avg_co)  else "—", "平均 CO 转化率",  "%")
        stat_card(c3, f"{avg_co2:.2f}" if not pd.isna(avg_co2) else "—", "平均 CO₂ 转化率", "%")
        stat_card(c4, f"{avg_h2co:.3f}" if not pd.isna(avg_h2co) else "—", "平均 H₂/CO", "")

        st.markdown("")

        # 展示列定义
        DISPLAY_COLS = {
            "catalyst_name_normalized": "催化剂名称",
            "_metal_primary":           "活性金属",
            "active_metal_loading_wt_pct": "活性金属负载量 wt%",
            "support":                  "载体",
            "promoter_elements":        "助剂元素",
            "BET_surface_area_m2_g":    "BET比表面积 m²/g",
            "metal_particle_size_nm":   "金属粒径 nm",
            "reaction_temperature_C":   "反应温度 °C",
            "reaction_pressure_bar":    "压力 bar",
            "H2_CO_ratio":              "H₂/CO 摩尔比",
            "CO_conversion_pct":        "CO 转化率 %",
            "CO2_conversion_pct":       "CO₂ 转化率 %",
            "C5plus_selectivity_pct":   "C₅⁺ 选择性 %",
            "CH4_selectivity_pct":      "CH₄ 选择性 %",
            "CO_or_CO2_route":          "原料气类型",
            "year":                     "年份",
            "paper_title":              "论文标题",
        }
        avail_cols = [c for c in DISPLAY_COLS if c in filtered.columns]
        show_df = filtered[avail_cols].rename(columns=DISPLAY_COLS).copy()
        if "原料气类型" in show_df.columns:
            show_df["原料气类型"] = "原料气"

        # None → 空字符串（美化显示）
        for c in show_df.columns:
            show_df[c] = show_df[c].where(show_df[c].notna(), other=None)

        st.dataframe(show_df, use_container_width=True, height=380,
                     column_config={
                         "BET比表面积 m²/g":   st.column_config.NumberColumn(format="%.1f"),
                         "金属粒径 nm":         st.column_config.NumberColumn(format="%.2f"),
                         "反应温度 °C":         st.column_config.NumberColumn(format="%.0f"),
                         "压力 bar":            st.column_config.NumberColumn(format="%.2f"),
                         "H₂/CO 摩尔比":        st.column_config.NumberColumn(format="%.2f"),
                         "CO 转化率 %":         st.column_config.NumberColumn(format="%.2f"),
                         "CO₂ 转化率 %":        st.column_config.NumberColumn(format="%.2f"),
                         "C₅⁺ 选择性 %":        st.column_config.NumberColumn(format="%.2f"),
                         "CH₄ 选择性 %":        st.column_config.NumberColumn(format="%.2f"),
                         "活性金属负载量 wt%":   st.column_config.NumberColumn(format="%.1f"),
                         "年份":                st.column_config.NumberColumn(format="%.0f"),
                     })

        # CSV 下载
        csv_bytes = show_df.to_csv(index=False).encode("utf-8-sig")
        st.download_button("⬇️ 下载筛选后的 CSV", csv_bytes,
                           f"FT_SAF_filtered_{n}records.csv", "text/csv",
                           use_container_width=True)

        # ── 散点图 ──
        st.markdown('<div class="section-title">筛选数据分布</div>', unsafe_allow_html=True)

        NUM_COLS = {
            "反应温度 °C":       "reaction_temperature_C",
            "压力 bar":          "reaction_pressure_bar",
            "H₂/CO 摩尔比":      "H2_CO_ratio",
            "CO 转化率 %":       "CO_conversion_pct",
            "CO₂ 转化率 %":      "CO2_conversion_pct",
            "C₅⁺ 选择性 %":      "C5plus_selectivity_pct",
            "CH₄ 选择性 %":      "CH4_selectivity_pct",
            "BET比表面积 m²/g":  "BET_surface_area_m2_g",
            "金属粒径 nm":       "metal_particle_size_nm",
            "活性金属负载量 wt%": "active_metal_loading_wt_pct",
            "年份":              "year",
        }
        avail_num = {k: v for k, v in NUM_COLS.items() if v in filtered.columns}

        sc1, sc2 = st.columns(2)
        with sc1:
            st.markdown("横轴")
            x_label = st.selectbox("", list(avail_num.keys()),
                                   index=list(avail_num.keys()).index("反应温度 °C") if "反应温度 °C" in avail_num else 0,
                                   key="db_x", label_visibility="collapsed")
        with sc2:
            st.markdown("纵轴")
            y_label = st.selectbox("", list(avail_num.keys()),
                                   index=list(avail_num.keys()).index("CO 转化率 %") if "CO 转化率 %" in avail_num else 1,
                                   key="db_y", label_visibility="collapsed")

        x_col = avail_num[x_label]
        y_col = avail_num[y_label]

        plot_df = filtered[[x_col, y_col, "_metal_primary", "catalyst_name_normalized",
                             "paper_title", "year"]].dropna(subset=[x_col, y_col]).copy()
        plot_df = plot_df.rename(columns={
            x_col: x_label, y_col: y_label,
            "_metal_primary": "活性金属",
            "catalyst_name_normalized": "催化剂",
            "paper_title": "论文",
        })

        if len(plot_df) > 0:
            fig_sc = px.scatter(
                plot_df, x=x_label, y=y_label,
                color="活性金属",
                hover_data=["催化剂", "论文", "year"],
                color_discrete_sequence=px.colors.qualitative.Set1,
                opacity=0.75,
                height=420,
            )
            fig_sc.update_traces(marker=dict(size=7))
            fig_sc.update_layout(
                margin=dict(t=10, b=10, l=10, r=10),
                legend=dict(title="活性金属", orientation="v"),
            )
            st.plotly_chart(fig_sc, use_container_width=True)
        else:
            st.info("当前筛选条件下无可绘图的数据，请放宽筛选范围。")


# ══════════════════════════════════════════════════════════════════
#  页面：论文自动下载（4步流程）
# ══════════════════════════════════════════════════════════════════
elif page == "📥 论文自动下载":

    # 用 session_state 管理当前步骤和数据
    if "dl_step" not in st.session_state:
        st.session_state["dl_step"] = 1
    if "clf_df"   not in st.session_state:
        st.session_state["clf_df"] = None
    if "dl_log"   not in st.session_state:
        st.session_state["dl_log"] = []
    if "manifest" not in st.session_state:
        st.session_state["manifest"] = []
    if "out_dir"  not in st.session_state:
        st.session_state["out_dir"] = "./downloads"
    if "webvpn_log" not in st.session_state:
        st.session_state["webvpn_log"] = pd.DataFrame()

    step = st.session_state["dl_step"]
    pipeline_bar(step)

    # ────────────────────────────────────────────────
    #  步骤1：DOI 分类
    # ────────────────────────────────────────────────
    st.markdown('<div class="section-title">① DOI 分类</div>', unsafe_allow_html=True)
    if True:

        col_up, col_tip = st.columns([2, 1])
        with col_up:
            uploaded = st.file_uploader("上传 xlsx / csv（含 DOI 列）",
                                        type=["csv","xlsx"], key="doi_upload",
                                        label_visibility="collapsed")
        with col_tip:
            st.markdown("""
            <div class="info-box">
            <b>识别的列名</b><br>
            <code>DOI</code> / <code>doi</code> — 必填<br>
            <code>access_type</code> — OA / non-OA<br>
            <code>title</code> — 文章标题<br>
            <code>year</code> — 发表年份
            </div>""", unsafe_allow_html=True)

        manual = st.text_area("或手动粘贴 DOI（每行一个）", height=90, key="manual_doi",
                              label_visibility="visible",
                              placeholder="10.1021/acscatal.4c00001\n10.1007/s10853-024-09012-1")

        if st.button("🔍 分析 DOI 列表", type="primary", use_container_width=True, key="btn_classify"):
            df_raw = None
            if uploaded:
                try:
                    df_raw = load_df(uploaded)
                except Exception as e:
                    st.error(f"文件读取失败：{e}")
            elif manual.strip():
                df_raw = pd.DataFrame({"doi": [d.strip() for d in manual.strip().splitlines() if d.strip()]})

            if df_raw is not None and not df_raw.empty:
                with st.spinner("识别出版商并分配路由…"):
                    result_df, ok = classify_df(df_raw)
                if ok:
                    st.session_state["clf_df"] = result_df
                    st.session_state["dl_step"] = 1  # 停留在步骤1展示结果

    # 步骤1结果展示
    if st.session_state["clf_df"] is not None:
        df = st.session_state["clf_df"]
        rc = Counter(df["路由"])
        total = len(df)
        n_dir = rc.get("direct",0) + rc.get("verified",0)
        n_vpn = rc.get("webvpn",0)
        n_rev = rc.get("review",0) + rc.get("probe",0)
        n_bad = rc.get("invalid",0)

        st.markdown('<div class="section-title">分类结果</div>', unsafe_allow_html=True)
        for col, val, lbl, desc in zip(st.columns(5),
            [total, n_dir, n_vpn, n_rev, n_bad],
            ["总计","可直接下载","需 WebVPN","待审核","无效"],
            ["全部文献","direct+verified","机构授权路由","probe+review","DOI缺失/无效"],
        ):
            stat_card(col, val, lbl, desc)

        st.markdown("")
        all_routes = sorted(df["路由"].unique())
        sel = st.multiselect("按路由筛选", all_routes, default=all_routes,
                             format_func=lambda r: ROUTE_LABEL.get(r, r), key="route_filter")
        df_show = df[df["路由"].isin(sel)] if sel else df
        show_cols = [c for c in ["doi","title","出版商","路由","下载URL","year"] if c in df_show.columns]
        st.dataframe(df_show[show_cols], use_container_width=True, height=300)

        st.markdown("")
        if st.button("下一步：开始下载 →", type="primary", key="goto_step2"):
            st.session_state["dl_step"] = 2
            st.rerun()

    # ────────────────────────────────────────────────
    #  步骤2：自动下载（直接下载 + WebVPN下载 两个子区块）
    # ────────────────────────────────────────────────
    if step >= 2 and st.session_state["clf_df"] is not None:
        st.markdown('<div class="section-title">② 自动下载</div>', unsafe_allow_html=True)
        if True:
            df = st.session_state["clf_df"]
            df_direct = df[df["路由"].isin(AUTO_ROUTES) & (df["下载URL"] != "")]
            df_webvpn = df[df["路由"] == "webvpn"]

            out_dir = st.text_input("PDF 保存目录", value=st.session_state["out_dir"], key="out_dir_input")
            st.session_state["out_dir"] = out_dir

            # ═══════════ 子区块A：直接下载（ACS/Springer/MDPI等，requests直连） ═══════════
            st.markdown("##### 🟢 直接下载（无需 WebVPN）")
            col_a, col_b = st.columns(2)
            with col_a:
                test_n = st.number_input("测试模式（0=全部）", 0, 50, 3, key="test_n")
            with col_b:
                min_kb = st.number_input("最小体积 KB", 5, 200, 20, key="min_kb")

            if len(df_direct) == 0:
                st.caption("当前队列中没有可直接下载的文献。")
            else:
                st.info(f"共 **{len(df_direct)}** 篇可直接下载 · {'测试前 '+str(test_n)+' 篇' if test_n else '全量模式'}")

            if st.button("🚀 开始直接下载", type="primary", use_container_width=True, key="btn_download"):
                queue = df_direct.to_dict("records")
                if test_n:
                    queue = queue[:test_n]

                prog  = st.progress(0)
                box   = st.empty()
                rows  = []
                sha_seen = {}

                for idx, row in enumerate(queue):
                    doi = str(row.get("doi",""))
                    pub = str(row.get("出版商",""))
                    url = str(row.get("下载URL",""))
                    delay = PUBLISHER_DELAY.get(pub, 2.0)

                    box.markdown(
                        f'<div class="info-box">⬇️ [{idx+1}/{len(queue)}] <b>{pub}</b> · <code>{doi}</code></div>',
                        unsafe_allow_html=True,
                    )
                    result = download_one(doi, url, pub, out_dir, int(min_kb))

                    if result.get("sha256") and result["status"] == "success":
                        if result["sha256"] in sha_seen:
                            result["status"] = "duplicate"
                            result["reason"] = f"与 {sha_seen[result['sha256']]} 内容重复"
                        else:
                            sha_seen[result["sha256"]] = doi

                    rows.append({
                        "doi": doi, "出版商": pub,
                        "status": result["status"],
                        "size_kb": result.get("size_kb",""),
                        "file": result.get("file",""),
                        "reason": result.get("reason",""),
                        "sha256": result.get("sha256",""),
                        "时间": datetime.now().strftime("%H:%M:%S"),
                    })
                    prog.progress((idx+1)/len(queue))
                    if idx < len(queue)-1:
                        time.sleep(delay)

                prog.empty(); box.empty()
                st.session_state["dl_log"]  = rows

            if st.session_state["dl_log"]:
                log = st.session_state["dl_log"]
                counts = Counter(r["status"] for r in log)
                st.markdown("**直接下载结果**")
                for col, k, lbl in zip(st.columns(6),
                    ["success","skipped_exists","no_access","likely_paywall","invalid_pdf","error"],
                    ["成功","已存在","无权限(403)","疑似付费墙","内容无效","错误"],
                ):
                    stat_card(col, counts.get(k,0), lbl)
                if counts.get("no_access",0) or counts.get("likely_paywall",0):
                    st.warning(
                        "⚠️ 存在无权限访问的文献（含明确403和疑似付费墙两种），"
                        "请确认已连接校园网或 VPN 后重试，或改用 WebVPN 下载这部分文献。"
                    )
                if counts.get("success",0):
                    st.success(f"✅ 成功下载 {counts['success']} 篇，保存至 {st.session_state['out_dir']}")
                st.dataframe(pd.DataFrame(log), use_container_width=True, height=220)

            st.markdown("---")

            # ═══════════ 子区块B：WebVPN下载（全部文献统一走WebVPN，不再按路由筛选） ═══════════
            st.markdown("##### 🟠 WebVPN 下载（处理全部上传文献，不区分路由分类）")
            st.caption(
                "WebVPN 可访问几乎所有出版商，因此这里不再按 direct/verified/webvpn 路由筛选，"
                "直接对你上传的全部文献依次尝试下载。"
            )

            # 云端部署检测：WebVPN依赖本地Chrome调试端口，云端无法使用
            _is_cloud = not any([
                Path("/home/claude").exists() and Path("C:/").exists(),  # 不像Windows本地
                os.environ.get("COMPUTERNAME"),  # Windows环境变量
                os.environ.get("USERPROFILE"),   # Windows用户目录
            ]) or os.environ.get("STREAMLIT_SERVER_HEADLESS") == "true"
            if not os.path.exists("C:\\") and not os.environ.get("COMPUTERNAME"):
                st.warning(
                    "⚠️ **云端部署环境检测**：WebVPN 下载功能依赖本地运行的 Chrome 浏览器（调试模式端口9222），"
                    "在 Streamlit Cloud 等云端环境中**无法使用**。\n\n"
                    "如需使用 WebVPN 下载，请将本项目下载到本地 Windows 电脑运行。"
                )

            chrome_online = check_chrome_debug_port()
            if chrome_online:
                st.success("✅ 已检测到调试模式 Chrome（端口9222在线）")
            else:
                st.error(
                    "❌ 未检测到调试模式 Chrome。请先：\n\n"
                    "1. 双击运行 `start_chrome_debug.bat`\n"
                    "2. 在弹出的Chrome窗口里手动登录 webvpn.xjtu.edu.cn\n"
                    "3. 保持该窗口开启，回到本页面点「🔃 重新检测」"
                )
                if st.button("🔃 重新检测 Chrome 状态", key="recheck_chrome"):
                    st.rerun()

            col_w1, col_w2 = st.columns(2)
            with col_w1:
                webvpn_script = st.text_input(
                    "webvpn_downloader3.py 路径",
                    value=str(Path(__file__).parent / "webvpn_downloader3.py"),
                    key="webvpn_script_path",
                )
            with col_w2:
                webvpn_delay = st.number_input("篇间延迟(秒)", 1.0, 10.0, 3.0, key="webvpn_delay")

            # 全部文献作为WebVPN处理队列，不再筛选路由
            df_webvpn_all = df.copy()
            total_count = len(df_webvpn_all)

            run_mode = st.radio(
                "处理范围",
                ["全部依次下载（推荐）", "先小批量测试"],
                horizontal=True,
                key="webvpn_run_mode",
            )

            if run_mode == "全部依次下载（推荐）":
                webvpn_limit = 0  # 0 表示不限制，脚本端会处理全部
                st.info(f"将对全部 **{total_count}** 篇文献依次发起下载，从第 1 篇开始处理到最后一篇。")
            else:
                webvpn_limit = st.number_input(
                    "测试篇数（从第1篇开始）", 1, max(total_count, 1),
                    min(10, total_count) if total_count else 1, key="webvpn_test_limit",
                )
                st.info(f"测试模式：将处理前 **{webvpn_limit}** 篇（共 {total_count} 篇）。")

            webvpn_disabled = (not chrome_online) or (total_count == 0) or (not Path(webvpn_script).exists())
            if not Path(webvpn_script).exists() and total_count > 0:
                st.warning(f"⚠️ 脚本路径不存在：`{webvpn_script}`，请确认 webvpn_downloader3.py 的实际位置。")

            if st.button("🚀 开始 WebVPN 下载", type="primary", use_container_width=True,
                        key="btn_webvpn_download", disabled=webvpn_disabled):
                queue_df = df_webvpn_all if webvpn_limit == 0 else df_webvpn_all.head(int(webvpn_limit))
                queue_csv_path = Path(out_dir) / "webvpn_queue_from_app.csv"
                write_webvpn_queue_csv(queue_df, queue_csv_path)

                status_box = st.empty()
                log_box = st.empty()
                status_box.info(f"⏳ 正在启动 WebVPN 下载（{len(queue_df)} 篇），实时日志如下…")

                # 记录运行前日志文件已有多少行，作为"本次运行"的基准线
                wv_log_path = Path(out_dir) / "webvpn_playwright_log.csv"
                rows_before = len(parse_webvpn_log(wv_log_path))

                proc = run_webvpn_downloader(
                    queue_csv_path, out_dir,
                    int(webvpn_limit), int(min_kb), 30, float(webvpn_delay),
                    webvpn_script,
                )

                output_lines = []
                for line in proc.stdout:
                    output_lines.append(line.rstrip())
                    log_box.code("\n".join(output_lines[-30:]))  # 只展示最近30行，避免页面过长

                proc.wait()
                status_box.success(f"✅ WebVPN 下载脚本已结束（退出码 {proc.returncode}）")

                # 只取本次运行新增的日志行，不包含之前历史运行的记录
                wv_log_df_full = parse_webvpn_log(wv_log_path)
                wv_log_df = wv_log_df_full.iloc[rows_before:].reset_index(drop=True) if not wv_log_df_full.empty else wv_log_df_full
                st.session_state["webvpn_log"] = wv_log_df

            if not st.session_state.get("webvpn_log", pd.DataFrame()).empty:
                wv_log_df = st.session_state["webvpn_log"]
                wv_counts = Counter(wv_log_df.get("status", pd.Series()).tolist())
                st.markdown("**WebVPN 下载结果（本次运行）**")
                for col, k, lbl in zip(st.columns(5),
                    ["success","no_pdf_found","invalid_pdf","timeout","error"],
                    ["成功","未找到PDF","内容无效","超时","错误"],
                ):
                    stat_card(col, wv_counts.get(k,0), lbl)
                st.dataframe(wv_log_df, use_container_width=True, height=240)

            st.markdown("---")

            if st.button("下一步：校验 PDF →", type="primary", key="goto_step3"):
                st.session_state["dl_step"] = 3
                st.rerun()

    # ────────────────────────────────────────────────
    #  步骤3：PDF 校验（完整集成 build_manifest.py）
    # ────────────────────────────────────────────────
    if step >= 3:
        st.markdown('<div class="section-title">③ PDF 校验</div>', unsafe_allow_html=True)
        if True:
            st.markdown("""
            <div class="info-box">
            完整集成 <code>build_manifest.py</code> 的校验逻辑，逐文件提取：<br>
            ① <code>%PDF-</code> 头部验证 &nbsp;
            ② 文件大小 &nbsp;
            ③ 页数读取 &nbsp;
            ④ 首页 DOI 提取 &nbsp;
            ⑤ PDF元数据标题/作者 &nbsp;
            ⑥ 加密检测 &nbsp;
            ⑦ SHA-256 去重 &nbsp;
            ⑧ 文件名解析（作者/年份/标题）
            </div>
            """, unsafe_allow_html=True)
            st.markdown("")

            col_dir, col_kb, col_sha = st.columns([3, 1, 1])
            with col_dir:
                scan_dir = st.text_input("扫描目录（下载 PDF 所在文件夹）",
                                         value=st.session_state.get("out_dir", "./downloads"),
                                         key="scan_dir")
            with col_kb:
                min_kb_scan = st.number_input("最小体积 KB", 5, 200, 20, key="min_kb_scan")
            with col_sha:
                do_sha = st.checkbox("计算 SHA-256", value=True, key="do_sha",
                                     help="大批量时可关闭以加快速度")

            col_scan1, col_scan2 = st.columns(2)
            with col_scan1:
                do_scan = st.button("🔍 扫描并校验 PDF", type="primary",
                                    use_container_width=True, key="btn_scan")
            with col_scan2:
                also_upload = st.file_uploader("或直接上传 PDF 文件校验",
                                               type=["pdf"], accept_multiple_files=True,
                                               key="pdf_upload", label_visibility="collapsed")

            # ── 扫描本地目录 ──
            if do_scan:
                folder = Path(scan_dir)
                if not folder.exists():
                    st.error(f"目录不存在：{scan_dir}")
                else:
                    with st.spinner(f"正在扫描 {scan_dir}，逐文件读取 PDF 元信息…"):
                        manifest = scan_pdf_folder(folder, int(min_kb_scan), bool(do_sha))
                    st.session_state["manifest"] = manifest

            # ── 直接上传 PDF 校验 ──
            if also_upload:
                import tempfile, os
                tmp_results = []
                sha_map: dict[str, str] = {}
                for uf in also_upload:
                    # 写到临时文件，复用 validate_pdf_file
                    data = uf.read()
                    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                        tmp.write(data)
                        tmp_path = Path(tmp.name)
                    try:
                        r = validate_pdf_file(tmp_path, int(min_kb_scan), bool(do_sha))
                        r["file_name"] = uf.name   # 恢复原始文件名
                        r["relative_path"] = uf.name
                    except Exception as e:
                        r = {"file_name": uf.name, "valid": False, "pdf_error": str(e)}
                    finally:
                        os.unlink(tmp_path)
                    # 去重
                    sha = r.get("sha256", "")
                    r["duplicate_of"] = sha_map.get(sha, "")
                    if sha and not r["duplicate_of"]:
                        sha_map[sha] = uf.name
                    tmp_results.append(r)
                st.session_state["manifest"] = tmp_results

            # ── 展示校验结果 ──
            if st.session_state.get("manifest"):
                manifest = st.session_state["manifest"]
                df_m = pd.DataFrame(manifest)

                total_m   = len(df_m)
                valid_m   = int(df_m["valid"].sum()) if "valid" in df_m.columns else 0
                invalid_m = total_m - valid_m
                dup_m     = int((df_m.get("duplicate_of","") != "").sum()) if "duplicate_of" in df_m.columns else 0
                enc_m     = int((df_m.get("is_encrypted", False) == True).sum()) if "is_encrypted" in df_m.columns else 0
                doi_found = int((df_m.get("first_page_doi","") != "").sum()) if "first_page_doi" in df_m.columns else 0

                st.markdown('<div class="section-title">校验结果</div>', unsafe_allow_html=True)
                for col, val, lbl, desc in zip(st.columns(6),
                    [total_m, valid_m, invalid_m, dup_m, enc_m, doi_found],
                    ["扫描文件", "有效 PDF", "无效/过小", "重复内容", "加密文件", "首页找到DOI"],
                    ["","头部+大小通过","","SHA-256相同","需解密","从首页文本提取"],
                ):
                    stat_card(col, val, lbl, desc)

                st.markdown("")

                # 选择展示列：优先展示最有用的字段
                priority_cols = [
                    "file_name", "file_size_kb", "page_count",
                    "pdf_header_valid", "pdf_read_ok", "valid",
                    "first_page_doi", "is_encrypted",
                    "pdf_metadata_title", "pdf_metadata_author",
                    "filename_year_guess", "filename_author_guess",
                    "sha256", "duplicate_of", "pdf_error", "relative_path",
                ]
                show_cols = [c for c in priority_cols if c in df_m.columns]
                st.dataframe(df_m[show_cols], use_container_width=True, height=340)

                # 问题汇总
                if invalid_m:
                    st.warning(f"⚠️ {invalid_m} 个文件校验不通过（头部非PDF或文件过小），可能下载不完整。")
                if dup_m:
                    st.warning(f"⚠️ {dup_m} 个文件内容重复（SHA-256相同），请检查是否误下载相同文章。")
                if enc_m:
                    st.info(f"ℹ️ {enc_m} 个文件已加密，后续 LLM 抽取前需先解密。")
                if valid_m == total_m and dup_m == 0:
                    st.success(f"✅ 全部 {total_m} 个 PDF 校验通过，无重复，质量良好！")

                if st.button("下一步：导出结果 →", type="primary", key="goto_step4"):
                    st.session_state["dl_step"] = 4
                    st.rerun()

    # ────────────────────────────────────────────────
    #  步骤4：导出结果
    # ────────────────────────────────────────────────
    if step >= 4:
        st.markdown('<div class="section-title">④ 导出结果</div>', unsafe_allow_html=True)
        if True:
            st.markdown('<div class="section-title">导出文件</div>', unsafe_allow_html=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M")
            df_clf = st.session_state.get("clf_df")

            c1, c2, c3, c4 = st.columns(4)

            # 完整分类队列
            with c1:
                if df_clf is not None:
                    st.download_button("⬇️ 完整 DOI 队列",
                        df_clf.to_csv(index=False).encode("utf-8-sig"),
                        f"queue_all_{ts}.csv", "text/csv", use_container_width=True)

            # WebVPN 队列（全部文献，与下载逻辑口径保持一致）
            with c2:
                if df_clf is not None:
                    st.download_button(f"⬇️ WebVPN 队列（全部 {len(df_clf)} 篇）",
                        df_clf.to_csv(index=False).encode("utf-8-sig"),
                        f"queue_webvpn_all_{ts}.csv", "text/csv", use_container_width=True)

            # 下载日志（直接下载 或 WebVPN下载，两者任一有数据就显示）
            with c3:
                has_direct_log = bool(st.session_state.get("dl_log"))
                has_webvpn_log = not st.session_state.get("webvpn_log", pd.DataFrame()).empty

                if has_direct_log and has_webvpn_log:
                    # 两种日志都有，合并导出
                    direct_df = pd.DataFrame(st.session_state["dl_log"])
                    direct_df["来源"] = "直接下载"
                    wv_df = st.session_state["webvpn_log"].copy()
                    wv_df["来源"] = "WebVPN下载"
                    combined_df = pd.concat([direct_df, wv_df], ignore_index=True, sort=False)
                    st.download_button("⬇️ 下载日志（直接+WebVPN）",
                        combined_df.to_csv(index=False).encode("utf-8-sig"),
                        f"download_log_{ts}.csv", "text/csv", use_container_width=True)
                elif has_direct_log:
                    log_df = pd.DataFrame(st.session_state["dl_log"])
                    st.download_button("⬇️ 下载日志（直接下载）",
                        log_df.to_csv(index=False).encode("utf-8-sig"),
                        f"download_log_{ts}.csv", "text/csv", use_container_width=True)
                elif has_webvpn_log:
                    log_df = st.session_state["webvpn_log"]
                    st.download_button("⬇️ 下载日志（WebVPN）",
                        log_df.to_csv(index=False).encode("utf-8-sig"),
                        f"download_log_{ts}.csv", "text/csv", use_container_width=True)
                else:
                    st.caption("暂无下载日志")

            # PDF 校验报告（manifest）
            with c4:
                if st.session_state.get("manifest"):
                    mf_df = pd.DataFrame(st.session_state["manifest"])
                    st.download_button("⬇️ PDF 校验报告",
                        mf_df.to_csv(index=False).encode("utf-8-sig"),
                        f"manifest_{ts}.csv", "text/csv", use_container_width=True)

            st.markdown("")
            st.markdown('<div class="section-title">本次运行汇总</div>', unsafe_allow_html=True)

            log = st.session_state.get("dl_log", [])
            mf  = st.session_state.get("manifest", [])
            wv_log = st.session_state.get("webvpn_log", pd.DataFrame())
            log_counts = Counter(r["status"] for r in log)
            wv_counts_summary = Counter(wv_log.get("status", pd.Series()).tolist()) if not wv_log.empty else Counter()
            mf_valid   = sum(1 for r in mf if r.get("valid"))

            total_success = log_counts.get("success",0) + wv_counts_summary.get("success",0)

            summary = {
                "运行时间":          datetime.now().strftime("%Y-%m-%d %H:%M"),
                "DOI总数":           len(df_clf) if df_clf is not None else 0,
                "直接下载成功":      log_counts.get("success",0) + log_counts.get("skipped_exists",0),
                "WebVPN下载成功":    wv_counts_summary.get("success",0),
                "下载成功合计":      total_success,
                "无权限(需VPN)":     log_counts.get("no_access",0),
                "PDF校验通过":       mf_valid,
                "PDF校验失败":       len(mf) - mf_valid,
            }
            for k, v in summary.items():
                st.markdown(
                    f'<div style="display:flex;justify-content:space-between;padding:7px 14px;'
                    f'background:#FAFAFA;border-radius:6px;margin-bottom:4px;font-size:0.85rem">'
                    f'<span style="color:#666">{k}</span><span style="font-weight:600">{v}</span></div>',
                    unsafe_allow_html=True,
                )

            st.markdown("")
            if st.button("↩️ 重新开始", key="btn_reset"):
                for k in ["dl_step","clf_df","dl_log","manifest"]:
                    st.session_state[k] = (1 if k=="dl_step" else ([] if k in ("dl_log","manifest") else None))
                st.session_state["webvpn_log"] = pd.DataFrame()
                st.rerun()


# ══════════════════════════════════════════════════════════════════
#  页面：自动抠取数据（4步流程：建索引 → 准备packet → 调用LLM → 质检）
# ══════════════════════════════════════════════════════════════════
elif page == "🔬 自动抠取数据":

    EXTRACTION_STEPS = ["① 建立PDF索引", "② 准备Prompt包", "③ LLM抽取", "④ 质检校验"]

    for k, v in [
        ("ex_step", 1), ("ex_manifest_csv", None), ("ex_packets_jsonl", None),
        ("ex_extractions_jsonl", None), ("ex_qc_report", None),
    ]:
        if k not in st.session_state:
            st.session_state[k] = v

    ex_step = st.session_state["ex_step"]
    pipeline_bar(ex_step, EXTRACTION_STEPS)

    # ========== 脚本目录和PDF根目录（相对路径，兼容云端和本地） ==========
    _base_dir = Path(__file__).parent
    DEFAULT_SCRIPTS_DIR = str(_base_dir)
    DEFAULT_PDF_ROOT    = str(_base_dir / "downloads")
    DEFAULT_OUT_DIR     = str(_base_dir / "outputs")
    # ================================================================

    # ────────────────────────────────────────
    #  步骤1：建立PDF索引（build_manifest.py）
    # ────────────────────────────────────────
    st.markdown('<div class="section-title">① 建立 PDF 索引</div>', unsafe_allow_html=True)
    if True:
        st.markdown("""
        <div class="info-box">
        扫描已下载的 PDF 目录（<code>OA/OA-paper</code> 和 <code>non_OA/non-OA</code> 子目录），
        提取文件名线索、页数、首页DOI，并尝试与已有 xlsx 元数据表匹配，生成统一索引表
        <code>manifest_full.csv</code>。这是 <code>build_manifest.py</code> 的功能。
        </div>
        """, unsafe_allow_html=True)

        col_s1, col_s2 = st.columns(2)
        with col_s1:
            scripts_dir = st.text_input("脚本所在目录", value=DEFAULT_SCRIPTS_DIR, key="ex_scripts_dir")
        with col_s2:
            pdf_root = st.text_input("PDF 根目录（含OA/non_OA子文件夹）", value=DEFAULT_PDF_ROOT, key="ex_pdf_root")

        out_dir = st.text_input("输出目录", value=DEFAULT_OUT_DIR, key="ex_out_dir")
        hash_files = st.checkbox("计算 SHA-256（用于去重，大批量时较慢）", value=False, key="ex_hash_files")

        if st.button("🔍 开始扫描并建立索引", type="primary", use_container_width=True, key="btn_build_manifest"):
            script_path = str(Path(scripts_dir) / "build_manifest.py")
            if not Path(script_path).exists():
                st.error(f"脚本不存在：{script_path}")
            else:
                # ======== 修改：将 "python" 改为 sys.executable ========
                cmd = [
                    sys.executable, script_path,
                    "--root", pdf_root,
                    "--out-dir", out_dir,
                    "--progress-every", "50",
                ]
                if hash_files:
                    cmd.append("--hash-files")
                st.markdown(f'<div class="cmd-box">{" ".join(cmd)}</div>', unsafe_allow_html=True)
                with st.spinner("正在扫描 PDF 并建立索引（视PDF数量可能需要几分钟）…"):
                    rc, out = run_cmd_sync(cmd, cwd=scripts_dir, timeout=1800)
                if rc == 0:
                    st.success("✅ 索引建立完成")
                    manifest_path = Path(out_dir) / "manifest_full.csv"
                    if manifest_path.exists():
                        st.session_state["ex_manifest_csv"] = str(manifest_path)
                else:
                    st.error(f"❌ 失败 (exit {rc})")
                st.code(out[-3000:] if len(out) > 3000 else out)

        if st.session_state["ex_manifest_csv"]:
            mf_path = Path(st.session_state["ex_manifest_csv"])
            if mf_path.exists():
                mf_df = read_log_tail(str(mf_path), 99999)
                if not mf_df.empty:
                    # 兼容列名
                    valid_col = None
                    for col in ["valid_pdf_gt20kb", "valid", "valid_pdf"]:
                        if col in mf_df.columns:
                            valid_col = col
                            break
                    if valid_col:
                        valid_n = int(mf_df[valid_col].astype(str).str.lower().isin(["true","1"]).sum())
                    elif "file_size_kb" in mf_df.columns:
                        valid_n = int((mf_df["file_size_kb"].astype(float) > 20).sum())
                    else:
                        valid_n = 0
                    st.markdown('<div class="section-title">索引结果</div>', unsafe_allow_html=True)
                    for col, val, lbl in zip(st.columns(3),
                        [len(mf_df), valid_n, len(mf_df)-valid_n],
                        ["扫描总数", "有效PDF", "无效/过小"]):
                        stat_card(col, val, lbl)
                    st.dataframe(mf_df.head(50), use_container_width=True, height=240)

                    if st.button("下一步：准备Prompt包 →", type="primary", key="ex_goto_step2"):
                        st.session_state["ex_step"] = 2
                        st.rerun()

    # ────────────────────────────────────────
    #  步骤2：准备Prompt包（prepare_prompt_packets.py）
    # ────────────────────────────────────────
    if ex_step >= 2 and st.session_state["ex_manifest_csv"]:
        st.markdown('<div class="section-title">② 准备 Prompt 包</div>', unsafe_allow_html=True)
        if True:
            st.markdown("""
            <div class="info-box">
            读取索引表，对每篇有效PDF按关键词在5个阶段（metadata / catalyst / reaction_conditions /
            performance / provenance_confidence）里分别挑选最相关的页面文字，打包成LLM可直接处理的
            任务包（JSONL，每行一个任务）。这是 <code>prepare_prompt_packets.py</code> 的功能。
            </div>
            """, unsafe_allow_html=True)

            col_p1, col_p2 = st.columns(2)
            with col_p1:
                pp_limit = st.number_input("处理论文数（0=全部）", 0, 5000, 5, key="ex_pp_limit")
            with col_p2:
                pp_stages = st.multiselect(
                    "处理阶段",
                    ["metadata", "catalyst", "reaction_conditions", "performance", "provenance_confidence"],
                    default=["metadata", "catalyst", "reaction_conditions", "performance", "provenance_confidence"],
                    key="ex_pp_stages",
                )

            if st.button("📦 生成 Prompt 包", type="primary", use_container_width=True, key="btn_prepare_packets"):
                script_path = str(Path(scripts_dir) / "prepare_prompt_packets.py")
                packets_out = str(Path(out_dir) / "prompt_packets.jsonl")
                if not Path(script_path).exists():
                    st.error(f"脚本不存在：{script_path}")
                else:
                    # ======== 修改：将 "python" 改为 sys.executable ========
                    cmd = [
                        sys.executable, script_path,
                        "--root", pdf_root,
                        "--manifest", st.session_state["ex_manifest_csv"],
                        "--out", packets_out,
                        "--stages", ",".join(pp_stages) if pp_stages else "all",
                        "--limit", str(pp_limit),
                    ]
                    st.markdown(f'<div class="cmd-box">{" ".join(cmd)}</div>', unsafe_allow_html=True)
                    with st.spinner("正在抽取页面文字并打包…"):
                        rc, out = run_cmd_sync(cmd, cwd=scripts_dir, timeout=1800)
                    if rc == 0:
                        st.success("✅ Prompt 包生成完成")
                        st.session_state["ex_packets_jsonl"] = packets_out
                    else:
                        st.error(f"❌ 失败 (exit {rc})")
                    st.code(out[-3000:] if len(out) > 3000 else out)

            if st.session_state["ex_packets_jsonl"] and Path(st.session_state["ex_packets_jsonl"]).exists():
                with open(st.session_state["ex_packets_jsonl"], encoding="utf-8") as f:
                    n_packets = sum(1 for _ in f)
                st.info(f"共生成 **{n_packets}** 条任务包，保存于 `{st.session_state['ex_packets_jsonl']}`")

                if st.button("下一步：LLM 抽取 →", type="primary", key="ex_goto_step3"):
                    st.session_state["ex_step"] = 3
                    st.rerun()

    # ────────────────────────────────────────
    #  步骤3：调用LLM抽取（run_llm_extraction.py）
    # ────────────────────────────────────────
    if ex_step >= 3 and st.session_state.get("ex_packets_jsonl"):
        st.markdown('<div class="section-title">③ LLM 数据抽取</div>', unsafe_allow_html=True)
        if True:
            st.markdown("""
            <div class="info-box">
            把每条 Prompt 包发给 LLM，对照 5 个阶段对应的提示词模板，拿回严格 JSON 格式的结构化数据。
            <b>API Key 请勿粘贴在群聊/文档里，建议优先用环境变量 <code>OPENAI_API_KEY</code> 提供。</b>
            </div>
            """, unsafe_allow_html=True)

            col_l1, col_l2 = st.columns(2)
            with col_l1:
                llm_model = st.text_input("模型名称", value="deepseek-v4-pro", key="ex_llm_model")
            with col_l2:
                llm_base_url = st.text_input("API Base URL", value="https://api.openai-proxy.org/v1", key="ex_llm_base_url")

            llm_api_key = st.text_input(
                "API Key（留空则使用环境变量 OPENAI_API_KEY）",
                value="", type="password", key="ex_llm_api_key",
                help="留空更安全：提前在系统里设置环境变量 OPENAI_API_KEY，这里就不用每次手填。"
            )

            col_l3, col_l4 = st.columns(2)
            with col_l3:
                llm_limit = st.number_input("处理任务数（0=全部）", 0, 99999, 10, key="ex_llm_limit")
            with col_l4:
                llm_delay = st.number_input("请求间隔(秒)", 0.0, 5.0, 0.5, key="ex_llm_delay")

            prompts_dir_input = st.text_input(
                "提示词模板目录", value=str(Path(scripts_dir) / "prompts"), key="ex_prompts_dir"
            )

            if st.button("🤖 开始 LLM 抽取", type="primary", use_container_width=True, key="btn_run_llm"):
                script_path = str(Path(scripts_dir) / "run_llm_extraction.py")
                extractions_out = str(Path(out_dir) / "extractions.jsonl")
                if not Path(script_path).exists():
                    st.error(f"脚本不存在：{script_path}")
                elif not llm_api_key and not os.environ.get("OPENAI_API_KEY"):
                    st.warning("⚠️ 未提供 API Key，且系统环境变量 OPENAI_API_KEY 也未设置，请填写后再试。")
                else:
                    # ======== 修改：将 "python" 改为 sys.executable ========
                    cmd = [
                        sys.executable, script_path,
                        "--packets", st.session_state["ex_packets_jsonl"],
                        "--prompts-dir", prompts_dir_input,
                        "--out", extractions_out,
                        "--model", llm_model,
                        "--base-url", llm_base_url,
                        "--limit", str(llm_limit),
                        "--delay", str(llm_delay),
                    ]
                    if llm_api_key:
                        cmd.extend(["--api-key", llm_api_key])

                    status_box = st.empty()
                    log_box = st.empty()
                    status_box.info("⏳ 正在调用 LLM 抽取，实时进度如下…")

                    proc = run_cmd_bg(cmd, cwd=scripts_dir)
                    output_lines = []
                    for line in proc.stdout:
                        output_lines.append(line.rstrip())
                        log_box.code("\n".join(output_lines[-30:]))
                    proc.wait()

                    if proc.returncode == 0:
                        status_box.success(f"✅ LLM 抽取完成")
                        st.session_state["ex_extractions_jsonl"] = extractions_out
                    else:
                        status_box.error(f"❌ 抽取脚本退出码 {proc.returncode}，请检查上方日志")

            if st.session_state.get("ex_extractions_jsonl") and Path(st.session_state["ex_extractions_jsonl"]).exists():
                with open(st.session_state["ex_extractions_jsonl"], encoding="utf-8") as f:
                    n_extracted = sum(1 for _ in f)
                st.info(f"共抽取 **{n_extracted}** 条结果，保存于 `{st.session_state['ex_extractions_jsonl']}`")

                if st.button("下一步：质检校验 →", type="primary", key="ex_goto_step4"):
                    st.session_state["ex_step"] = 4
                    st.rerun()

    # ────────────────────────────────────────
    #  步骤4：质检校验（qc_validate_extractions.py）+ 导出CSV
    # ────────────────────────────────────────
    if ex_step >= 4 and st.session_state.get("ex_extractions_jsonl"):
        st.markdown('<div class="section-title">④ 质检校验</div>', unsafe_allow_html=True)
        if True:
            st.markdown("""
            <div class="info-box">
            检查每条LLM抽取结果的JSON结构是否合规：必填字段是否齐全、confidence是否在0-1范围、
            数值字段是否带单位、是否每个paper都覆盖了全部5个阶段。这是
            <code>qc_validate_extractions.py</code> 的功能。
            </div>
            """, unsafe_allow_html=True)

            if st.button("✅ 开始质检", type="primary", use_container_width=True, key="btn_run_qc"):
                script_path = str(Path(scripts_dir) / "qc_validate_extractions.py")
                qc_report_path = str(Path(out_dir) / "qc_report.json")
                if not Path(script_path).exists():
                    st.error(f"脚本不存在：{script_path}")
                else:
                    # ======== 修改：将 "python" 改为 sys.executable ========
                    cmd = [sys.executable, script_path, st.session_state["ex_extractions_jsonl"], "--out", qc_report_path]
                    st.markdown(f'<div class="cmd-box">{" ".join(cmd)}</div>', unsafe_allow_html=True)
                    with st.spinner("正在校验…"):
                        rc, out = run_cmd_sync(cmd, cwd=scripts_dir, timeout=300)
                    if rc == 0:
                        st.success("✅ 质检完成")
                        st.session_state["ex_qc_report"] = qc_report_path
                    else:
                        st.error(f"❌ 失败 (exit {rc})")

            if st.session_state.get("ex_qc_report") and Path(st.session_state["ex_qc_report"]).exists():
                qc_data = json.loads(Path(st.session_state["ex_qc_report"]).read_text(encoding="utf-8"))

                st.markdown('<div class="section-title">质检结果</div>', unsafe_allow_html=True)
                for col, val, lbl in zip(st.columns(4),
                    [qc_data.get("parsed_records",0), qc_data.get("papers_seen",0),
                     qc_data.get("missing_stage_count",0),
                     qc_data.get("severity_counts",{}).get("error",0)],
                    ["解析记录数", "覆盖论文数", "缺失阶段的论文", "错误数"]):
                    stat_card(col, val, lbl)

                if qc_data.get("issue_counts"):
                    st.markdown("**问题分布**")
                    issue_df = pd.DataFrame(
                        list(qc_data["issue_counts"].items()), columns=["问题类型", "出现次数"]
                    ).sort_values("出现次数", ascending=False)
                    st.dataframe(issue_df, use_container_width=True, height=200)

                if qc_data.get("missing_stage_examples"):
                    st.markdown("**缺失阶段示例（最多50条）**")
                    st.dataframe(pd.DataFrame(qc_data["missing_stage_examples"]), use_container_width=True, height=200)

                st.download_button(
                    "⬇️ 导出完整质检报告",
                    json.dumps(qc_data, ensure_ascii=False, indent=2).encode("utf-8"),
                    "qc_report.json", "application/json",
                )

                if st.session_state.get("ex_extractions_jsonl"):
                    extractions_bytes = Path(st.session_state["ex_extractions_jsonl"]).read_bytes()
                    st.download_button(
                        "⬇️ 导出全部抽取结果（JSONL）",
                        extractions_bytes,
                        "extractions.jsonl", "application/json",
                    )

            # ── 新增：导出为 CSV ──
            st.markdown("---")
            st.markdown("#### 📊 导出抽取结果为 CSV")
            jsonl_path = st.session_state.get("ex_extractions_jsonl")
            if jsonl_path and Path(jsonl_path).exists():
                if st.button("🔄 生成 CSV 表格", key="btn_gen_csv"):
                    with st.spinner("正在转换 JSONL 为 CSV..."):
                        # 读取 JSONL
                        records = []
                        with open(jsonl_path, "r", encoding="utf-8") as f:
                            for line in f:
                                if line.strip():
                                    records.append(json.loads(line))
                        if records:
                            # 展平并合并为长表
                            flat_rows = []
                            for rec in records:
                                base = {"paper_id": rec["paper_id"], "stage": rec["stage"]}
                                if rec["stage"] == "metadata":
                                    meta = rec.get("metadata", {})
                                    for k, v in meta.items():
                                        if isinstance(v, dict):
                                            base[f"metadata_{k}"] = v.get("value")
                                            base[f"metadata_{k}_confidence"] = v.get("confidence")
                                        else:
                                            base[f"metadata_{k}"] = v
                                    flat_rows.append(base)
                                elif rec["stage"] == "catalyst":
                                    catalysts = rec.get("catalysts", [])
                                    for cat in catalysts:
                                        row = base.copy()
                                        # 处理可能为dict或直接值的字段
                                        def get_val(d, key):
                                            val = d.get(key) if isinstance(d, dict) else None
                                            if isinstance(val, dict):
                                                return val.get("value")
                                            return val
                                        row["catalyst_id"] = get_val(cat, "catalyst_id")
                                        row["catalyst_name"] = get_val(cat, "catalyst_name_or_code")
                                        row["active_metal"] = ", ".join([m.get("value") for m in cat.get("active_metal", []) if isinstance(m, dict) and m.get("value")]) if cat.get("active_metal") else None
                                        row["support"] = get_val(cat, "support")
                                        row["promoter"] = ", ".join([p.get("value") for p in cat.get("promoter", []) if isinstance(p, dict) and p.get("value")]) if cat.get("promoter") else None
                                        row["preparation_method"] = get_val(cat, "preparation_method")
                                        reduction = cat.get("reduction", {})
                                        if isinstance(reduction, dict):
                                            row["reduction_temp"] = reduction.get("temperature")
                                            row["reduction_time"] = reduction.get("time")
                                            row["reduction_atmosphere"] = reduction.get("atmosphere")
                                        flat_rows.append(row)
                                elif rec["stage"] == "reaction_conditions":
                                    conditions = rec.get("reaction_conditions", [])
                                    for cond in conditions:
                                        row = base.copy()
                                        row["condition_id"] = cond.get("condition_id")
                                        row["catalyst_ref"] = get_val(cond, "catalyst_ref")
                                        row["temperature"] = get_val(cond, "temperature")
                                        row["pressure"] = get_val(cond, "pressure")
                                        row["space_velocity"] = get_val(cond, "space_velocity")
                                        row["h2_co_ratio"] = get_val(cond, "h2_co_ratio")
                                        row["reactor"] = get_val(cond, "reactor")
                                        row["time_on_stream"] = get_val(cond, "time_on_stream")
                                        flat_rows.append(row)
                                elif rec["stage"] == "performance":
                                    performances = rec.get("performance_records", [])
                                    for perf in performances:
                                        row = base.copy()
                                        row["performance_id"] = perf.get("performance_id")
                                        row["catalyst_ref"] = get_val(perf, "catalyst_ref")
                                        row["metric"] = perf.get("metric")
                                        row["species_or_range"] = perf.get("species_or_range")
                                        row["value"] = perf.get("value")
                                        row["unit"] = perf.get("unit")
                                        row["confidence"] = perf.get("confidence")
                                        flat_rows.append(row)
                                elif rec["stage"] == "provenance_confidence":
                                    provs = rec.get("provenance_records", [])
                                    for prov in provs:
                                        row = base.copy()
                                        row["target_stage"] = prov.get("target_stage")
                                        row["target_field"] = prov.get("target_field")
                                        row["page"] = prov.get("page")
                                        row["table_or_figure"] = prov.get("table_or_figure")
                                        row["caption"] = prov.get("caption")[:200] + "..." if prov.get("caption") and len(prov["caption"]) > 200 else prov.get("caption")
                                        row["evidence_text"] = prov.get("evidence_text")[:200] + "..." if prov.get("evidence_text") and len(prov["evidence_text"]) > 200 else prov.get("evidence_text")
                                        row["confidence"] = prov.get("confidence")
                                        flat_rows.append(row)
                            if flat_rows:
                                df_csv = pd.DataFrame(flat_rows)
                                csv_data = df_csv.to_csv(index=False).encode("utf-8-sig")

                                # ── 宽表转换：每篇论文一行，所有字段平铺 ──
                                meta_rows = df_csv[df_csv["stage"]=="metadata"].drop(columns=["stage"], errors="ignore")
                                cat_rows  = df_csv[df_csv["stage"]=="catalyst"].drop(columns=["stage"], errors="ignore")
                                cond_rows = df_csv[df_csv["stage"]=="reaction_conditions"].drop(columns=["stage"], errors="ignore")
                                perf_rows = df_csv[df_csv["stage"]=="performance"].drop(columns=["stage"], errors="ignore")
                                prov_rows = df_csv[df_csv["stage"]=="provenance_confidence"].drop(columns=["stage"], errors="ignore")

                                # metadata：每篇一行，直接用
                                wide = meta_rows.copy()

                                # catalyst/conditions/performance：每篇可能多行，
                                # 取第一条，列名加前缀区分
                                def first_row_per_paper(sub_df, prefix):
                                    if sub_df.empty:
                                        return pd.DataFrame()
                                    first = sub_df.groupby("paper_id").first().reset_index()
                                    rename_map = {c: f"{prefix}_{c}" for c in first.columns if c != "paper_id"}
                                    return first.rename(columns=rename_map)

                                stage_groups = [
                                    (cat_rows,  "cat",  "Catalyst（催化剂）"),
                                    (cond_rows, "cond", "Reaction Conditions（反应条件）"),
                                    (perf_rows, "perf", "Performance（性能指标）"),
                                    (prov_rows, "prov", "Provenance & Confidence（溯源置信度）"),
                                ]
                                for sub, prefix, _ in stage_groups:
                                    sub_wide = first_row_per_paper(sub, prefix)
                                    if not sub_wide.empty:
                                        wide = wide.merge(sub_wide, on="paper_id", how="left")

                                # ── 各阶段数据量提示 ──
                                stage_info = {
                                    "Metadata":               len(meta_rows),
                                    "Catalyst":               len(cat_rows),
                                    "Reaction Conditions":    len(cond_rows),
                                    "Performance":            len(perf_rows),
                                    "Provenance & Confidence": len(prov_rows),
                                }
                                info_parts = [f"{k}：**{v}** 条" if v > 0 else f"~~{k}：0条（LLM未返回）~~" for k, v in stage_info.items()]
                                st.info("各阶段数据量 → " + " | ".join(info_parts))

                                # ── 生成带分组表头的 Excel 文件 ──
                                import io as _io
                                try:
                                    import openpyxl
                                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                                    excel_buf = _io.BytesIO()
                                    wb = openpyxl.Workbook()
                                    ws = wb.active
                                    ws.title = "宽表"

                                    # 构建列→大类映射
                                    col_group_map = {}
                                    for col in wide.columns:
                                        if col == "paper_id":
                                            col_group_map[col] = "基本信息"
                                        elif col.startswith("cat_"):
                                            col_group_map[col] = "Catalyst（催化剂）"
                                        elif col.startswith("cond_"):
                                            col_group_map[col] = "Reaction Conditions（反应条件）"
                                        elif col.startswith("perf_"):
                                            col_group_map[col] = "Performance（性能指标）"
                                        elif col.startswith("prov_"):
                                            col_group_map[col] = "Provenance & Confidence（溯源置信度）"
                                        else:
                                            col_group_map[col] = "Metadata（论文信息）"

                                    # 配色方案
                                    group_colors = {
                                        "基本信息":                              "D6E4F0",
                                        "Metadata（论文信息）":                   "D0E8FF",
                                        "Catalyst（催化剂）":                     "D5F5E3",
                                        "Reaction Conditions（反应条件）":        "FFF3CD",
                                        "Performance（性能指标）":                "FADBD8",
                                        "Provenance & Confidence（溯源置信度）":  "EDE7F6",
                                    }

                                    cols = list(wide.columns)
                                    # 第1行：大类组标题（合并同组单元格）
                                    groups_ordered = []
                                    prev_group = None
                                    group_start = 1
                                    for i, col in enumerate(cols, 1):
                                        grp = col_group_map[col]
                                        if grp != prev_group:
                                            if prev_group is not None:
                                                groups_ordered.append((prev_group, group_start, i-1))
                                            prev_group = grp
                                            group_start = i
                                    groups_ordered.append((prev_group, group_start, len(cols)))

                                    for grp, start, end in groups_ordered:
                                        cell = ws.cell(row=1, column=start, value=grp)
                                        fill_color = group_colors.get(grp, "FFFFFF")
                                        cell.fill = PatternFill("solid", fgColor=fill_color)
                                        cell.font = Font(bold=True, size=11)
                                        cell.alignment = Alignment(horizontal="center", vertical="center")
                                        if end > start:
                                            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)

                                    # 第2行：字段名
                                    for i, col in enumerate(cols, 1):
                                        cell = ws.cell(row=2, column=i, value=col)
                                        fill_color = group_colors.get(col_group_map[col], "FFFFFF")
                                        cell.fill = PatternFill("solid", fgColor=fill_color)
                                        cell.font = Font(bold=True, size=10)
                                        cell.alignment = Alignment(horizontal="center")

                                    # 数据行
                                    for row_idx, (_, row_data) in enumerate(wide.iterrows(), 3):
                                        for col_idx, col in enumerate(cols, 1):
                                            val = row_data[col]
                                            try:
                                                is_na = pd.isna(val)
                                                if hasattr(is_na, '__iter__'):
                                                    is_na = True
                                            except Exception:
                                                is_na = False
                                            ws.cell(row=row_idx, column=col_idx, value="" if is_na else str(val))

                                    # 自动列宽
                                    for col_idx, col in enumerate(cols, 1):
                                        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, min(30, len(str(col))+2))

                                    ws.row_dimensions[1].height = 22
                                    ws.row_dimensions[2].height = 18
                                    ws.freeze_panes = "B3"

                                    wb.save(excel_buf)
                                    excel_bytes = excel_buf.getvalue()
                                    has_excel = True
                                except ImportError:
                                    has_excel = False

                                wide_data = wide.to_csv(index=False).encode("utf-8-sig")

                                col_dl1, col_dl2, col_dl3 = st.columns(3)
                                with col_dl1:
                                    st.download_button(
                                        label="⬇️ 长表 CSV（每阶段一行）",
                                        data=csv_data,
                                        file_name=f"extractions_long_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                                        mime="text/csv",
                                    )
                                with col_dl2:
                                    st.download_button(
                                        label="⬇️ 宽表 CSV（每篇论文一行）",
                                        data=wide_data,
                                        file_name=f"extractions_wide_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                                        mime="text/csv",
                                    )
                                with col_dl3:
                                    if has_excel:
                                        st.download_button(
                                            label="⬇️ 宽表 Excel（带分组表头）",
                                            data=excel_bytes,
                                            file_name=f"extractions_wide_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        )
                                    else:
                                        st.caption("安装 `openpyxl` 后可导出带分组表头的 Excel")
                                st.success(f"✅ 长表 {len(flat_rows)} 行 / 宽表 {len(wide)} 行（每篇论文一行）")
                            else:
                                st.warning("未找到任何有效记录")
                        else:
                            st.warning("JSONL 文件为空或格式不正确")
            else:
                st.info("请先完成 LLM 抽取，生成 extractions.jsonl 文件。")

            st.markdown("")
            if st.button("↩️ 重新开始抽取流程", key="ex_btn_reset"):
                for k in ["ex_step", "ex_manifest_csv", "ex_packets_jsonl", "ex_extractions_jsonl", "ex_qc_report"]:
                    st.session_state[k] = 1 if k == "ex_step" else None
                st.rerun()

# ══════════════════════════════════════════════════════════════════
#  页面：数据可视化分析（箱线图 / 折线图 / 散点图 / 云雨图）
# ══════════════════════════════════════════════════════════════════
elif page == "📈 数据可视化分析":

    _base_dir = Path(__file__).parent
    # 自动搜索可能的CSV位置，包括带时间戳的文件名
    def _find_extractions_csv(base):
        candidates = [
            base / "data" / "extractions_long.csv",
            base / "outputs" / "extractions_long.csv",
            base / "extractions_long.csv",
            Path(r"D:\agent\saf_extraction\outputs\extractions_long.csv"),
        ]
        # 还要搜索 data/ 下面带时间戳的文件，如 extractions_long_20260621_1608.csv
        for folder in [base / "data", base / "outputs", base]:
            if folder.exists():
                for f in sorted(folder.glob("extractions_long*.csv"), reverse=True):
                    candidates.insert(0, f)
        return next((str(p) for p in candidates if p.exists()), str(base / "data" / "extractions_long.csv"))

    DEFAULT_EXTRACTIONS_CSV = _find_extractions_csv(_base_dir)

    st.markdown("""
    <div class="info-box">
    读取第二阶段生成的结构化抽取结果（长表CSV），自动识别并统一单位
    （温度→°C，压力→MPa，时长→h），再提供4种图表用于探索数据分布与趋势。
    </div>
    """, unsafe_allow_html=True)

    col_u1, col_u2 = st.columns([3, 1])
    with col_u1:
        csv_path_input = st.text_input("抽取结果 CSV 路径", value=DEFAULT_EXTRACTIONS_CSV, key="viz_csv_path")
    with col_u2:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        load_clicked = st.button("📂 加载数据", type="primary", use_container_width=True, key="viz_load_btn")

    uploaded_csv = st.file_uploader("或者直接上传 CSV 文件", type=["csv"], key="viz_csv_upload",
                                    label_visibility="collapsed")

    if "viz_stages" not in st.session_state:
        st.session_state["viz_stages"] = None

    df_raw = None
    if uploaded_csv is not None:
        df_raw = pd.read_csv(uploaded_csv, encoding="utf-8-sig")
    elif load_clicked:
        p = Path(csv_path_input)
        if not p.exists():
            st.error(f"文件不存在：{p}")
        else:
            df_raw = pd.read_csv(p, encoding="utf-8-sig")

    if df_raw is not None:
        with st.spinner("正在清洗数据、识别并统一单位…"):
            stages = split_stages(df_raw)
        st.session_state["viz_stages"] = stages
        st.session_state["viz_raw_rowcount"] = len(df_raw)
        st.success(f"✅ 已加载 {len(df_raw)} 行，覆盖 {df_raw['paper_id'].nunique()} 篇论文")

    stages = st.session_state.get("viz_stages")

    if stages is None:
        st.info("请先加载数据。")
    else:
        meta_df = stages.get("metadata", pd.DataFrame())
        cond_df = stages.get("reaction_conditions", pd.DataFrame())
        perf_df = stages.get("performance", pd.DataFrame())
        cat_df  = stages.get("catalyst", pd.DataFrame())

        plot_perf_df = build_plot_ready_performance(perf_df, meta_df)
        plot_cond_df = build_plot_ready_conditions(cond_df, meta_df)
        plot_meta_df = build_plot_ready_metadata(meta_df)
        plot_cat_df  = build_plot_ready_catalyst(cat_df, meta_df)
        unparseable_df = get_unparseable_rows(stages)

        # ── 数据质量摘要 ──
        st.markdown('<div class="section-title">数据质量摘要</div>', unsafe_allow_html=True)
        for col, val, lbl in zip(st.columns(5),
            [meta_df["paper_id"].nunique() if "paper_id" in meta_df.columns else 0,
             len(perf_df), len(plot_perf_df), len(plot_cond_df), len(unparseable_df)],
            ["覆盖论文数", "性能记录总数", "可画图的性能数值", "可画图的反应条件数值", "无法解析为数值"]):
            stat_card(col, val, lbl)

        if not unparseable_df.empty:
            st.markdown('<div class="section-title">⚠️ 无法解析为数值的原始记录（不会进入下方图表）</div>', unsafe_allow_html=True)
            if True:
                st.dataframe(unparseable_df, use_container_width=True, height=200)

        # ── 指标大类总览：固定展示04_performance.md定义的5大类+其他，即使某类当前无数据 ──
        if not plot_perf_df.empty:
            st.markdown('<div class="section-title">性能指标大类总览（来自 04_performance.md）</div>', unsafe_allow_html=True)
            category_order = [cat for cat, _ in METRIC_CATEGORY_RULES] + [METRIC_CATEGORY_OTHER]
            category_counts = plot_perf_df["metric"].value_counts().reindex(category_order, fill_value=0)
            cat_cols = st.columns(len(category_order))
            for col, cat in zip(cat_cols, category_order):
                count = int(category_counts[cat])
                col.metric(cat, count, help="该大类下当前可画图的数值记录条数" if count else "当前数据中暂无该类指标的有效数值")
            st.caption(
                "提示：某大类显示0，通常是因为论文报告了该指标但LLM未能从原文提取出具体数值"
                "（例如只说'显著提升'而没给出数字），不代表归并规则遗漏了该类别。"
            )

        st.markdown("---")

        if plot_perf_df.empty and plot_cond_df.empty:
            st.warning("当前数据中没有可用于画图的数值记录，请检查抽取结果或加载更多论文。")
        else:
            data_source = st.radio(
                "数据来源",
                ["性能指标（转化率/选择性/产率等）",
                 "反应条件（温度/压力/运行时长）",
                 "论文信息（年份/期刊/工艺路线）",
                 "催化剂/材料（活性金属/载体/助剂）"],
                horizontal=True, key="viz_data_source",
            )

            # 路由到对应的数据集
            if data_source.startswith("性能指标"):
                active_df = plot_perf_df
            elif data_source.startswith("反应条件"):
                active_df = plot_cond_df
            elif data_source.startswith("论文信息"):
                active_df = pd.DataFrame()   # 论文信息用专属UI，不走通用active_df路径
            else:
                active_df = pd.DataFrame()   # 催化剂/材料用专属UI

            tab_box, tab_line, tab_scatter, tab_rain = st.tabs(
                ["📦 箱线图", "📈 折线图", "🔵 散点图", "☔ 云雨图"]
            )

            if not active_df.empty:
                present_metrics = set(active_df["metric"].dropna().unique().tolist())
                if data_source.startswith("性能指标"):
                    # 性能指标：按04_performance.md定义的大类顺序排列，结构清晰
                    ordered = [cat for cat, _ in METRIC_CATEGORY_RULES] + [METRIC_CATEGORY_OTHER]
                    metric_options = [m for m in ordered if m in present_metrics]
                    if len(metric_options) < len(ordered):
                        missing = [m for m in ordered if m not in present_metrics]
                        st.caption(
                            f"💡 当前可选 {len(metric_options)}/{len(ordered)} 个大类"
                            f"（{'、'.join(missing)} 暂无可用数值，详见上方「指标大类总览」）"
                        )
                else:
                    metric_options = sorted(present_metrics)
            else:
                metric_options = []

            # ════════════════════ 箱线图 ════════════════════
            with tab_box:
                st.caption(f"按所选「{data_source}」分组，展示数值分布的中位数、四分位距与离群点。")
                col_b1, col_b2 = st.columns(2)
                with col_b1:
                    box_metrics = st.multiselect("选择指标", metric_options,
                        default=metric_options[:5] if metric_options else [], key="box_metrics",
                        help="只列出当前数据中「有实际数值」的大类。某大类（如yield/STY/productivity）"
                             "未出现在此列表，通常是因为论文提到了该指标但未给出具体数字，"
                             "并非系统不支持——上方「指标大类总览」会展示全部6个大类的真实数据量。")
                with col_b2:
                    group_options = ["metric", "species_or_range", "metadata_journal"]
                    if "metric_raw" in active_df.columns:
                        group_options.insert(1, "metric_raw")
                    box_group_by = st.selectbox("分组方式（metric=归并大类，metric_raw=论文原文写法）",
                        group_options, key="box_group_by")

                box_data = active_df[active_df["metric"].isin(box_metrics)] if box_metrics else active_df
                if box_data.empty:
                    st.info("请选择至少一个指标。")
                else:
                    fig_box = px.box(
                        box_data, x=box_group_by, y="value_numeric", points="all",
                        color=box_group_by,
                        labels={"value_numeric": "数值", box_group_by: box_group_by},
                    )
                    fig_box.update_layout(showlegend=False, height=480)
                    plotly_chart_with_doi(fig_box, box_data if not box_data.empty else active_df, chart_key="doi_box")

            # ════════════════════ 折线图 ════════════════════
            with tab_line:
                st.caption(f"按年份展示「{data_source}」随时间的变化趋势（取每年均值）。")
                if active_df.empty or "metadata_year" not in active_df.columns or active_df["metadata_year"].dropna().empty:
                    st.info("数据中缺少年份信息，无法绘制折线图。")
                else:
                    line_metrics = st.multiselect("选择指标", metric_options,
                        default=metric_options[:3] if metric_options else [], key="line_metrics",
                        help="只列出当前数据中「有实际数值」的大类，原因同箱线图——"
                             "完整的6大类体系请看上方「指标大类总览」。")
                    line_data = active_df[active_df["metric"].isin(line_metrics)] if line_metrics else active_df.iloc[0:0]
                    if line_data.empty:
                        st.info("请选择至少一个指标。")
                    else:
                        trend = (
                            line_data.dropna(subset=["metadata_year"])
                            .groupby(["metadata_year", "metric"])["value_numeric"]
                            .mean()
                            .reset_index()
                        )
                        fig_line = px.line(
                            trend, x="metadata_year", y="value_numeric", color="metric",
                            markers=True, labels={"value_numeric": "均值", "metadata_year": "年份"},
                        )
                        fig_line.update_layout(height=480)
                        plotly_chart_with_doi(fig_line, line_data if not line_data.empty else active_df, chart_key="doi_line")

            # ════════════════════ 散点图 ════════════════════
            with tab_scatter:
                scatter_mode = st.radio(
                    "对比方式",
                    [f"同一数据源内部对比（{data_source}）", "反应条件 vs 性能指标（如：温度对转化率的影响）"],
                    key="scatter_mode",
                )

                if scatter_mode.startswith("同一数据源"):
                    st.caption("查看同一来源下两个指标之间的关系，或单一指标按置信度着色的分布。")
                    col_s1, col_s2, col_s3 = st.columns(3)
                    with col_s1:
                        scatter_metric_x = st.selectbox("X轴指标", metric_options, key="scatter_x") if metric_options else None
                    with col_s2:
                        scatter_metric_y = st.selectbox("Y轴指标（可选，留空则用论文序号）",
                            ["(无，按论文序号展开)"] + metric_options, key="scatter_y")
                    with col_s3:
                        color_options = [c for c in ["confidence", "metadata_journal", "species_or_range"] if c in active_df.columns]
                        color_by = st.selectbox("着色依据", color_options, key="scatter_color") if color_options else None

                    if scatter_metric_x:
                        x_data = active_df[active_df["metric"] == scatter_metric_x].copy()
                        merged = None
                        hover_cols_x = ["paper_id"] + (["metric_raw"] if "metric_raw" in x_data.columns else [])
                        if scatter_metric_y == "(无，按论文序号展开)" or not scatter_metric_y:
                            x_data = x_data.reset_index(drop=True)
                            x_data["序号"] = x_data.index
                            fig_scatter = px.scatter(
                                x_data, x="序号", y="value_numeric", color=color_by,
                                hover_data=hover_cols_x,
                                labels={"value_numeric": scatter_metric_x},
                            )
                        else:
                            y_cols = ["paper_id", "value_numeric"] + (["metric_raw"] if "metric_raw" in active_df.columns else [])
                            y_data = active_df[active_df["metric"] == scatter_metric_y][y_cols].rename(
                                columns={"value_numeric": "y_value", "metric_raw": "y_metric_raw"}
                            )
                            merged = x_data.merge(y_data, on="paper_id", how="inner")
                            if merged.empty:
                                st.info("两个指标在同一篇论文里没有同时出现的数据点。")
                            else:
                                hover_cols_merge = ["paper_id"]
                                if "metric_raw" in merged.columns:
                                    hover_cols_merge.append("metric_raw")
                                if "y_metric_raw" in merged.columns:
                                    hover_cols_merge.append("y_metric_raw")
                                fig_scatter = px.scatter(
                                    merged, x="value_numeric", y="y_value", color=color_by,
                                    hover_data=hover_cols_merge,
                                    labels={"value_numeric": scatter_metric_x, "y_value": scatter_metric_y},
                                )
                        if scatter_metric_y == "(无，按论文序号展开)" or (merged is not None and not merged.empty):
                            fig_scatter.update_layout(height=480)
                            plotly_chart_with_doi(fig_scatter, x_data if not x_data.empty else active_df, chart_key="doi_scatter_same")
                    else:
                        st.info("当前没有可选指标。")

                else:
                    st.caption("把反应条件（温度/压力/运行时长）和性能指标按「同一论文+同一催化剂」关联，查看条件对性能的影响。")
                    joined_df = join_conditions_with_performance(plot_cond_df, plot_perf_df)
                    if joined_df.empty:
                        st.info(
                            "没有能关联起来的数据：可能是反应条件和性能记录的 catalyst_ref 命名不一致，"
                            "或当前数据量太少。可以在下方「查看清洗后的各阶段数据表」里检查 catalyst_ref 字段。"
                        )
                    else:
                        col_c1, col_c2 = st.columns(2)
                        condition_col_names = [label for label, unit in CONDITION_FIELD_LABELS.values()]
                        condition_cols = [c for c in condition_col_names if c in joined_df.columns]
                        with col_c1:
                            cond_field = st.selectbox("反应条件（X轴）", condition_cols, key="join_cond_field") if condition_cols else None
                        with col_c2:
                            perf_metrics_in_join = sorted(joined_df["metric"].dropna().unique().tolist())
                            perf_field = st.selectbox("性能指标（Y轴）", perf_metrics_in_join, key="join_perf_field") if perf_metrics_in_join else None

                        if cond_field and perf_field:
                            plot_data = joined_df[joined_df["metric"] == perf_field].dropna(subset=[cond_field, "value_numeric"])
                            if plot_data.empty:
                                st.info("该组合下没有同时具备两个数值的数据点。")
                            else:
                                hover_cols_join = ["paper_id"] + (["metric_raw"] if "metric_raw" in plot_data.columns else []) + (["doi"] if "doi" in plot_data.columns else [])
                                show_trendline = _HAS_STATSMODELS and len(plot_data) >= 3
                                fig_join = px.scatter(
                                    plot_data, x=cond_field, y="value_numeric",
                                    color="catalyst_ref", hover_data=hover_cols_join,
                                    labels={cond_field: cond_field, "value_numeric": perf_field},
                                    trendline="ols" if show_trendline else None,
                                )
                                fig_join.update_layout(height=480)
                                if "doi" in plot_data.columns and plot_data["doi"].notna().any():
                                    scatter_with_doi_links(fig_join, plot_data, doi_col="doi", height=480, chart_key="join_scatter_doi")
                                else:
                                    st.plotly_chart(fig_join, use_container_width=True)
                                if not _HAS_STATSMODELS and len(plot_data) >= 3:
                                    st.caption(
                                        "💡 安装 `statsmodels` 后可在此图上显示趋势线（回归线）："
                                        "`pip install statsmodels`"
                                    )
                        else:
                            st.info("当前关联数据中缺少可选的反应条件或性能指标。")

            # ════════════════════ 云雨图 (Raincloud Plot) ════════════════════
            with tab_rain:
                st.caption(
                    f"云雨图（数据来源：{data_source}）= 分布形状(小提琴) + 抖动散点(原始数据点) + 箱线图(统计摘要)，"
                    "三层叠加，比单独箱线图更能看出数据真实分布形态。"
                )
                rain_metrics = st.multiselect("选择指标（建议选2-6个，太多会拥挤）", metric_options,
                    default=metric_options[:4] if metric_options else [], key="rain_metrics",
                    help="只列出当前数据中「有实际数值」的大类，原因同箱线图——"
                         "完整的6大类体系请看上方「指标大类总览」。")

                rain_data = active_df[active_df["metric"].isin(rain_metrics)] if rain_metrics else active_df.iloc[0:0]

                if rain_data.empty:
                    st.info("请选择至少一个指标。")
                else:
                    fig_rain = go.Figure()
                    colors = px.colors.qualitative.Set2
                    for i, m in enumerate(rain_metrics):
                        sub = rain_data[rain_data["metric"] == m]
                        color = colors[i % len(colors)]

                        # 第一层：小提琴（只画一侧，形成"云"的形状），偏移到分类右侧
                        fig_rain.add_trace(go.Violin(
                            x=[m] * len(sub), y=sub["value_numeric"],
                            side="positive", width=1.8, points=False,
                            line_color=color, fillcolor=color, opacity=0.45,
                            showlegend=False, scalemode="width",
                        ))
                        # 第二层：抖动散点（"雨滴"），叠在小提琴左侧
                        jitter = pd.Series(sub.index).apply(lambda _: __import__("random").uniform(-0.25, -0.05))
                        fig_rain.add_trace(go.Scatter(
                            x=[m] * len(sub), y=sub["value_numeric"],
                            mode="markers",
                            marker=dict(color=color, size=6, opacity=0.6),
                            showlegend=False,
                        ))
                        # 第三层：箱线图（统计摘要），细窄叠在中间
                        fig_rain.add_trace(go.Box(
                            x=[m] * len(sub), y=sub["value_numeric"],
                            width=0.15, line_color="black", fillcolor="rgba(255,255,255,0.7)",
                            showlegend=False, boxpoints=False,
                        ))

                    fig_rain.update_layout(
                        height=520, violingap=0.3, violinmode="overlay",
                        yaxis_title="数值", xaxis_title="指标",
                    )
                    plotly_chart_with_doi(fig_rain, rain_data if not rain_data.empty else active_df, chart_key="doi_rain")

        st.markdown("---")

        # ════════ 论文信息专属图表区 ════════
        if data_source.startswith("论文信息"):
            if plot_meta_df.empty:
                st.info("暂无论文信息数据。")
            else:
                st.markdown('<div class="section-title">论文信息分析</div>', unsafe_allow_html=True)
                meta_tab1, meta_tab2, meta_tab3 = st.tabs(["📅 年份分布", "🗺️ 工艺路线分布", "📰 期刊分布"])

                with meta_tab1:
                    st.caption("每年发表的论文数量趋势")
                    if "year" in plot_meta_df.columns:
                        year_counts = (
                            plot_meta_df.dropna(subset=["year"])
                            .groupby("year")["paper_id"].count()
                            .reset_index()
                            .rename(columns={"paper_id": "论文数量", "year": "年份"})
                        )
                        fig_year = px.bar(year_counts, x="年份", y="论文数量", color_discrete_sequence=["#1565C0"])
                        fig_year.update_layout(height=420)
                        plotly_chart_with_doi(fig_year, plot_meta_df, chart_key="doi_meta_year_bar", doi_col="doi")

                        # 折线趋势叠加
                        fig_line_year = px.line(year_counts, x="年份", y="论文数量",
                            markers=True, color_discrete_sequence=["#1976D2"])
                        fig_line_year.update_layout(height=380, title="年份趋势折线图")
                        plotly_chart_with_doi(fig_line_year, plot_meta_df, chart_key="doi_meta_year_line", doi_col="doi")
                    else:
                        st.info("数据中缺少年份字段。")

                with meta_tab2:
                    st.caption("各工艺路线（route_type）的论文数量分布")
                    if "route_type" in plot_meta_df.columns:
                        route_counts = (
                            plot_meta_df.dropna(subset=["route_type"])
                            .groupby("route_type")["paper_id"].count()
                            .reset_index()
                            .rename(columns={"paper_id": "论文数量", "route_type": "工艺路线"})
                            .sort_values("论文数量", ascending=False)
                        )
                        fig_route = px.bar(route_counts, x="工艺路线", y="论文数量",
                            color_discrete_sequence=["#1565C0"])
                        fig_route.update_layout(height=420)
                        plotly_chart_with_doi(fig_route, plot_meta_df, chart_key="doi_meta_route_bar", doi_col="doi")

                        # 工艺路线随年份变化
                        if "year" in plot_meta_df.columns:
                            route_year = (
                                plot_meta_df.dropna(subset=["route_type","year"])
                                .groupby(["year","route_type"])["paper_id"].count()
                                .reset_index()
                                .rename(columns={"paper_id":"论文数量","year":"年份","route_type":"工艺路线"})
                            )
                            fig_route_year = px.line(route_year, x="年份", y="论文数量",
                                color="工艺路线", markers=True)
                            fig_route_year.update_layout(height=420, title="各工艺路线随年份变化趋势")
                            plotly_chart_with_doi(fig_route_year, plot_meta_df, chart_key="doi_meta_route_year", doi_col="doi")
                    else:
                        st.info("数据中缺少工艺路线字段。")

                with meta_tab3:
                    st.caption("发文数量最多的期刊")
                    if "journal" in plot_meta_df.columns:
                        journal_counts = (
                            plot_meta_df.dropna(subset=["journal"])
                            .groupby("journal")["paper_id"].count()
                            .reset_index()
                            .rename(columns={"paper_id":"论文数量","journal":"期刊"})
                            .sort_values("论文数量", ascending=False)
                            .head(20)
                        )
                        fig_journal = px.bar(journal_counts, x="论文数量", y="期刊",
                            orientation="h", color_discrete_sequence=["#1565C0"])
                        fig_journal.update_layout(height=max(380, len(journal_counts)*28))
                        plotly_chart_with_doi(fig_journal, plot_meta_df, chart_key="doi_meta_journal", doi_col="doi")
                    else:
                        st.info("数据中缺少期刊字段。")

        # ════════ 催化剂/材料：整合进tabs（与性能指标/反应条件统一入口）════════
        elif data_source.startswith("催化剂"):
            if plot_cat_df.empty:
                st.info("暂无催化剂/材料数据。")
            else:
                cat_fields_avail = sorted(plot_cat_df["metric"].dropna().unique().tolist())
                cat_tab_bar, cat_tab_box, cat_tab_trend = st.tabs(
                    ["📊 材料频次条形图", "📦 材料分组箱线图（数值指标）", "📅 材料随年份变化"]
                )

                with cat_tab_bar:
                    st.caption("各类材料字段的出现频次分布")
                    sel_fields = st.multiselect("选择字段", cat_fields_avail,
                        default=cat_fields_avail, key="cat_bar_fields")
                    if sel_fields:
                        top_n = st.slider("每个字段最多展示前N种", 5, 30, 15, key="cat_top_n")
                        for i_field, field in enumerate(sel_fields):
                            field_data = plot_cat_df[plot_cat_df["metric"]==field]["value_text"].value_counts().head(top_n).reset_index()
                            field_data.columns = ["材料", "论文数"]
                            fig_cat = px.bar(field_data, x="材料", y="论文数",
                                title=f"{field} 分布（Top {top_n}）",
                                color_discrete_sequence=["#1565C0"])
                            fig_cat.update_layout(height=380)
                            sub_cat_df = plot_cat_df[plot_cat_df["metric"]==field].copy()
                            plotly_chart_with_doi(fig_cat, sub_cat_df, chart_key=f"doi_cat_bar_{i_field}", doi_col="doi")

                with cat_tab_box:
                    st.caption(
                        "把还原温度、还原时长等数值字段，按'活性金属'或'载体'分组画箱线图，"
                        "这样可以对比不同材料体系的参数分布差异。"
                    )
                    # 提取数值字段，同时关联活性金属/载体信息用于分组
                    numeric_rows2 = []
                    for _, row in cat_df.iterrows():
                        pid = row.get("paper_id","")
                        metal = str(row.get("active_metal","") or "").strip()
                        support = str(row.get("support","") or "").strip()
                        for col_name, label in [("reduction_temp","还原温度"), ("reduction_time","还原时长")]:
                            val = row.get(col_name)
                            if val is not None and str(val).strip() not in ("","nan","None"):
                                try:
                                    num_val = float(re.sub(r"[^\d.\-]","",str(val)) or "nan")
                                    if not pd.isna(num_val):
                                        numeric_rows2.append({
                                            "paper_id": pid,
                                            "指标": label,
                                            "数值": num_val,
                                            "活性金属": metal or "未知",
                                            "载体": support or "未知",
                                        })
                                except Exception:
                                    pass

                    if numeric_rows2:
                        num_df2 = pd.DataFrame(numeric_rows2)
                        group_by = st.selectbox("按什么分组", ["活性金属", "载体", "指标"], key="cat_box_group")
                        sel_num_field = st.selectbox("数值指标", num_df2["指标"].unique().tolist(), key="cat_box_field")
                        sub_num = num_df2[num_df2["指标"]==sel_num_field]
                        if not sub_num.empty:
                            fig_cat_box = px.box(
                                sub_num, x=group_by, y="数值", points="all",
                                color=group_by,
                                title=f"{sel_num_field} 按{group_by}分组分布",
                                labels={"数值": sel_num_field},
                            )
                            fig_cat_box.update_layout(height=460, showlegend=False)
                            # 关联DOI
                            sub_num_with_doi = sub_num.copy()
                            if "paper_id" in sub_num_with_doi.columns and not meta_df.empty and "metadata_doi" in meta_df.columns:
                                doi_map = meta_df.set_index("paper_id")["metadata_doi"].astype(str)
                                sub_num_with_doi["doi"] = sub_num_with_doi["paper_id"].map(doi_map)
                            plotly_chart_with_doi(fig_cat_box, sub_num_with_doi, chart_key="doi_cat_box", doi_col="doi")
                        else:
                            st.info("该数值指标暂无数据。")
                    else:
                        st.info(
                            "当前数据中暂无可解析的数值型催化剂参数（reduction_temp 等）。"
                            "全量数据跑完后会有更多。"
                        )

                with cat_tab_trend:
                    st.caption("各类材料随年份的论文数量变化")
                    if "metadata_year" in plot_cat_df.columns:
                        sel_field_year = st.selectbox("选择字段", cat_fields_avail, key="cat_year_field")
                        sel_values = st.multiselect(
                            "选择具体材料（留空=全部前15）",
                            plot_cat_df[plot_cat_df["metric"]==sel_field_year]["value_text"].value_counts().head(15).index.tolist(),
                            key="cat_year_values"
                        )
                        cat_year_data = plot_cat_df[
                            (plot_cat_df["metric"]==sel_field_year) & plot_cat_df["metadata_year"].notna()
                        ]
                        if sel_values:
                            cat_year_data = cat_year_data[cat_year_data["value_text"].isin(sel_values)]
                        if not cat_year_data.empty:
                            trend = (
                                cat_year_data.groupby(["metadata_year","value_text"])["paper_id"]
                                .count().reset_index()
                                .rename(columns={"paper_id":"论文数","metadata_year":"年份","value_text":sel_field_year})
                            )
                            fig_cat_year = px.line(trend, x="年份", y="论文数",
                                color=sel_field_year, markers=True)
                            fig_cat_year.update_layout(height=450)
                            plotly_chart_with_doi(fig_cat_year, cat_year_data, chart_key="doi_cat_trend", doi_col="doi")
                        else:
                            st.info("暂无可用数据。")
                    else:
                        st.info("数据中缺少年份信息。")

        st.markdown('<div class="section-title">📋 查看清洗后的各阶段数据表</div>', unsafe_allow_html=True)
        if True:
            sub_tabs = st.tabs(["metadata", "catalyst", "reaction_conditions", "performance"])
            for tab, name, data in zip(sub_tabs,
                ["metadata", "catalyst", "reaction_conditions", "performance"],
                [meta_df, cat_df, cond_df, perf_df]):
                with tab:
                    if data.empty:
                        st.caption("无数据")
                    else:
                        st.dataframe(data, use_container_width=True, height=300)

# ══════════════════════════════════════════════════════════════════
#  页面：使用说明
# ══════════════════════════════════════════════════════════════════
elif page == "📖 使用说明":

    # ── 文件路径信息框 ──
    st.markdown("""
<style>
.doc-paths {
    background: #f8faff;
    border: 1px solid #BBDEFB;
    border-radius: 8px;
    padding: 14px 20px;
    margin-bottom: 22px;
    font-family: 'Times New Roman','Microsoft YaHei','微软雅黑',serif;
    font-size: 0.9rem;
    line-height: 2.0;
    color: #1A237E;
}
.doc-paths code {
    background: #E3F2FD;
    color: #1565C0;
    padding: 1px 6px;
    border-radius: 3px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.82rem;
}
.doc-section-title {
    font-size: 1.12rem;
    font-weight: 700;
    color: #0D2B5E;
    border-left: 4px solid #1565C0;
    padding-left: 10px;
    margin: 24px 0 12px;
    font-family: 'Times New Roman','Microsoft YaHei','微软雅黑',serif;
}
.doc-module-list {
    font-family: 'Times New Roman','Microsoft YaHei','微软雅黑',serif;
    font-size: 0.95rem;
    line-height: 2.0;
    color: #222;
    padding-left: 4px;
}
.doc-module-list li {
    margin-bottom: 6px;
}
.doc-module-list b {
    color: #0D2B5E;
}
.doc-boundary {
    font-family: 'Times New Roman','Microsoft YaHei','微软雅黑',serif;
    font-size: 0.95rem;
    line-height: 1.85;
    color: #333;
    margin-top: 6px;
}
</style>
""", unsafe_allow_html=True)

    st.markdown(f"""
<div class="doc-paths">
    <div><b>文献数据目录：</b><code>data/FT_SAF_catalyst_extraction_wide_table.xlsx</code>（放在项目根目录的 data/ 子目录下）</div>
    <div><b>PDF 存放目录：</b><code>downloads/</code>（运行时自动创建于项目根目录）</div>
    <div><b>抽取结果文件：</b><code>outputs/extractions_long.csv</code>（运行时自动创建于项目根目录）</div>
</div>
""", unsafe_allow_html=True)

    # ── 页面模块说明 ──
    st.markdown('<div class="doc-section-title">页面模块</div>', unsafe_allow_html=True)
    st.markdown("""
<ol class="doc-module-list">
  <li><b>平台总览</b>：展示文献库规模（总量、有效 PDF、年份范围、OA 比例）、出版商分布以及自动下载效率对比。</li>
  <li><b>数据库浏览</b>：按催化剂家族、活性金属、原料气类型、年份、反应温度等条件筛选全量抽取结果，支持散点图可视化和导出 CSV。</li>
  <li><b>论文自动下载</b>：上传含 DOI 的 CSV / Excel，系统按出版商路由自动分类，支持一键批量下载并导出下载结果清单。</li>
  <li><b>自动抠取数据</b>：对已下载的 PDF 文件执行多阶段结构化抽取，输出催化剂组成、反应条件与催化性能数据，结果可导出为 CSV。</li>
  <li><b>数据可视化分析</b>：加载抽取结果，提供多维度图表分析，包括年份趋势、出版商分布、催化剂参数分布、性能散点图及分组箱线图。</li>
  <li><b>使用说明</b>：说明各模块功能、数据文件位置及生产应用边界（本页）。</li>
</ol>
""", unsafe_allow_html=True)

    # ── 应用边界 ──
    st.markdown('<div class="doc-section-title">应用边界</div>', unsafe_allow_html=True)
    st.markdown("""
<div class="doc-boundary">
当前系统是文献数据驱动的科研辅助平台。模型可以帮助快速筛选候选催化剂和观察趋势，但不能直接替代实验验证。
正式用于企业项目时，需要进行数据人工复核、外部验证、适用域判断和不确定性评估。<br><br>
本平台仅使用机构授权 / 合规公开渠道获取文献，不绕过付费墙或访问控制。
所有下载行为均遵循出版商服务条款，下载频率已设置合理延迟以避免对服务器造成压力。
</div>
""", unsafe_allow_html=True)

    # ── 注意事项 ──
    st.markdown('<div class="doc-section-title">注意事项</div>', unsafe_allow_html=True)
    st.markdown("""
<ol class="doc-module-list">
  <li>上传的 CSV / Excel 文件须包含 <code>DOI</code> 列，列名大小写不限。</li>
  <li>自动抠取数据模块依赖大语言模型 API，请确保网络连通且 API Key 已正确配置。</li>
  <li>数据可视化分析需先完成数据抽取并上传对应的 JSON / CSV 结果文件。</li>
  <li>平台统计数字（1,627 篇等）来自当前已扫描的数据集，随数据更新会同步变化。</li>
  <li>如遇下载失败，可查看导出的结果清单中的"状态"列，手动处理标注为 <code>🟠 需 WebVPN</code> 或 <code>🔴 人工审核</code> 的条目。</li>
</ol>
""", unsafe_allow_html=True)
